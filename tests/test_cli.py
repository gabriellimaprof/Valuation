"""Testes da linha de comando."""

from __future__ import annotations

import pytest
from openpyxl import load_workbook

from valuation.cli import main
from valuation.exemplos import escrever_exemplos


@pytest.fixture
def exemplos(tmp_path):
    empresa, comparaveis = escrever_exemplos(tmp_path)
    return str(empresa), str(comparaveis)


def test_comando_exemplo_gera_arquivos(tmp_path, capsys):
    destino = tmp_path / "novos"
    assert main(["exemplo", "--destino", str(destino)]) == 0
    assert (destino / "empresa_exemplo.yaml").exists()
    assert (destino / "comparaveis_exemplo.yaml").exists()


def test_comando_wacc(exemplos, capsys):
    assert main(["wacc", exemplos[0]]) == 0
    saida = capsys.readouterr().out
    assert "WACC (BRL nominal)" in saida
    assert "Beta realavancado" in saida


def test_comando_dcf(exemplos, capsys):
    assert main(["dcf", exemplos[0]]) == 0
    saida = capsys.readouterr().out
    assert "Enterprise Value" in saida
    assert "Equity Value" in saida
    assert "FCFF (fluxo para a firma)" in saida


def test_comando_dcf_com_meio_de_ano_vale_mais(exemplos, capsys):
    main(["dcf", exemplos[0]])
    fim_de_ano = capsys.readouterr().out
    main(["dcf", exemplos[0], "--meio-de-ano"])
    meio_de_ano = capsys.readouterr().out

    def equity(texto: str) -> float:
        linha = next(l for l in texto.splitlines() if "Equity Value." in l)
        return float(linha.split()[-3].replace(",", ""))

    assert equity(meio_de_ano) > equity(fim_de_ano)


def test_comando_dcf_gera_excel(exemplos, tmp_path):
    destino = tmp_path / "saida" / "modelo.xlsx"
    assert main(["dcf", exemplos[0], "--excel", str(destino)]) == 0
    assert destino.exists()

    wb = load_workbook(destino)
    for aba in ("Leia-me", "Premissas", "Custo de Capital", "Projecao", "DCF"):
        assert aba in wb.sheetnames
    # O exemplo declara sensibilidade, cenarios e simulacao.
    for aba in ("Sensibilidade", "Cenarios", "Monte Carlo"):
        assert aba in wb.sheetnames


def test_comando_dcf_com_comparaveis(exemplos, capsys):
    assert main(["dcf", exemplos[0], "--comparaveis", exemplos[1]]) == 0
    saida = capsys.readouterr().out
    assert "Avaliacao relativa" in saida
    assert "EV/EBITDA" in saida


def test_comando_multiplos(exemplos, capsys):
    assert main(["multiplos", exemplos[1]]) == 0
    saida = capsys.readouterr().out
    assert "Estatisticas do peer group" in saida
    assert "Valor implicito" in saida


def test_comando_multiplos_com_referencia(exemplos, capsys):
    assert main(["multiplos", exemplos[1], "--referencia", "3o quartil"]) == 0
    assert "3o quartil do setor" in capsys.readouterr().out


def test_arquivo_inexistente_devolve_codigo_de_erro(capsys):
    assert main(["dcf", "nao_existe.yaml"]) == 1
    assert "Erro:" in capsys.readouterr().err


def test_premissa_invalida_devolve_codigo_de_erro(tmp_path, capsys):
    caminho = tmp_path / "ruim.yaml"
    caminho.write_text("nome: X\nmacro:\n  aliquota_ir: 34\n", encoding="utf-8")
    assert main(["wacc", str(caminho)]) == 1
    assert "decimais" in capsys.readouterr().err


def test_monte_carlo_sem_bloco_avisa(tmp_path, capsys):
    caminho = tmp_path / "sem_sim.yaml"
    caminho.write_text(
        "nome: X\n"
        "operacionais:\n"
        "  receita_base: 1000.0\n"
        "  horizonte: 3\n"
        "  crescimento_receita: 0.05\n"
        "  margem_ebitda: 0.20\n"
        "  depreciacao_pct_receita: 0.05\n"
        "  capex_pct_receita: 0.05\n"
        "  capital_giro_pct_receita: 0.10\n",
        encoding="utf-8",
    )
    assert main(["dcf", str(caminho), "--monte-carlo"]) == 0
    assert "nao declara o bloco" in capsys.readouterr().out


def test_cenario_base_reproduz_o_numero_principal(exemplos, capsys):
    """O cenario Base do exemplo tem que bater com o resultado do topo do relatorio."""
    main(["dcf", exemplos[0], "--meio-de-ano"])
    saida = capsys.readouterr().out

    linha_ev = next(l for l in saida.splitlines() if "Enterprise Value." in l)
    ev_principal = float(linha_ev.split()[-3].replace(",", ""))

    linha_cenarios = next(
        l for l in saida.splitlines() if l.strip().startswith("enterprise_value")
    )
    ev_base = float(linha_cenarios.split()[2].replace(",", ""))

    assert ev_base == pytest.approx(ev_principal, rel=1e-3)

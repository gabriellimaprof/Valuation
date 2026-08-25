"""Testes do leitor dos Dados Abertos da CVM.

Os fixtures em ``tests/dados/cvm`` nao foram escritos a mao: sao recortes em
bytes dos arquivos publicados em dados.cvm.gov.br -- o zip da DFP de 2023 e
2024 e o cadastro de companhias --, filtrados para quatro companhias e com o
latin-1, o ``;`` e o CRLF originais preservados. Testar contra planilha
inventada foi exatamente o que deixou passar os bugs que este projeto ja pagou
para descobrir.

As quatro companhias do recorte cobrem, cada uma, um comportamento diferente do
arquivo real:

======================  =====  ==========================================
Companhia               CVM    Por que esta aqui
======================  =====  ==========================================
WEG S.A.                 5410  ESCALA_MOEDA = MIL; publica con e ind
Vivara Participacoes    24805  ESCALA_MOEDA = UNIDADE (a pegadinha)
Sao Martinho            20516  exercicio social fecha em 31/03
Elektro Redes           17485  so publica demonstracao individual
======================  =====  ==========================================
"""

from __future__ import annotations

import zipfile
from pathlib import Path

import numpy as np
import pytest

from valuation.historico import analisar, sugerir_premissas
from valuation.importacao import (
    POR_CHAVE,
    Demonstracoes,
    aplicar_mapeamento_manual,
    carregar_abas,
)
from valuation.importacao.cvm import (
    ENCODING,
    SEPARADOR,
    Companhia,
    ErroCVM,
    _apenas_da_companhia,
    _fator_escala,
    baixar_dfp,
    buscar_companhias,
    carregar_cadastro,
    importar_cvm,
)

DADOS = Path(__file__).parent / "dados" / "cvm"

WEG, VIVARA, SAO_MARTINHO, ELEKTRO, BANCO_BRASIL = 5410, 24805, 20516, 17485, 1023

# Contas cujo valor o app move de lugar de proposito: elas nao batem com a linha
# publicada, e ha teste especifico para cada reclassificacao.
_RECLASSIFICADAS = {
    "variacao_capital_giro",
    "fluxo_operacional",
    "fluxo_investimento",
    "fluxo_financiamento",
    "capex",
    "arrendamento_curto_prazo",
    "arrendamento_longo_prazo",
    "divida_curto_prazo",
    "divida_longo_prazo",
}


@pytest.fixture(scope="module")
def catalogo() -> list[Companhia]:
    return carregar_cadastro(DADOS / "cad_cia_aberta.csv")


@pytest.fixture(scope="module")
def weg() -> Demonstracoes:
    return importar_cvm(WEG, [2023, 2024], cache=DADOS)


# ---------------------------------------------------------------------------
# O contrato do arquivo: o que foi conferido no portal, virado teste
# ---------------------------------------------------------------------------


def test_arquivo_da_cvm_e_latin1_e_nao_utf8():
    """O encoding e latin-1. Ler como UTF-8 estoura na primeira acentuacao.

    Fixa a descoberta para que ninguem "padronize" o modulo para UTF-8 por
    parecer o certo: o arquivo da CVM nao e UTF-8, e a excecao abaixo prova.
    """
    with zipfile.ZipFile(DADOS / "dfp_cia_aberta_2024.zip") as arquivo:
        bruto = arquivo.read("dfp_cia_aberta_DRE_con_2024.csv")

    with pytest.raises(UnicodeDecodeError):
        bruto.decode("utf-8")

    texto = bruto.decode(ENCODING)
    assert "Demonstração do Resultado" in texto
    assert "PENÚLTIMO" in texto


def test_layout_das_colunas_da_dfp():
    """As colunas que o leitor consome existem, com o separador esperado."""
    esperado_bp = [
        "CNPJ_CIA", "DT_REFER", "VERSAO", "DENOM_CIA", "CD_CVM", "GRUPO_DFP",
        "MOEDA", "ESCALA_MOEDA", "ORDEM_EXERC", "DT_FIM_EXERC", "CD_CONTA",
        "DS_CONTA", "VL_CONTA", "ST_CONTA_FIXA",
    ]
    with zipfile.ZipFile(DADOS / "dfp_cia_aberta_2024.zip") as arquivo:
        cabecalho_bp = (
            arquivo.read("dfp_cia_aberta_BPA_con_2024.csv")
            .split(b"\r\n")[0].decode(ENCODING).split(SEPARADOR)
        )
        cabecalho_dre = (
            arquivo.read("dfp_cia_aberta_DRE_con_2024.csv")
            .split(b"\r\n")[0].decode(ENCODING).split(SEPARADOR)
        )

    assert cabecalho_bp == esperado_bp
    # A DRE e a DFC descrevem um periodo, nao uma data: tem DT_INI_EXERC a mais.
    assert cabecalho_dre == esperado_bp[:9] + ["DT_INI_EXERC"] + esperado_bp[9:]


def test_nenhum_campo_contem_o_separador():
    """O recorte em bytes depende disso: CD_CVM sempre na mesma posicao.

    Se a CVM passar a citar campos com ';' dentro, a contagem de campos varia e
    o filtro rapido silenciosamente deixaria de achar a empresa. Melhor descobrir
    aqui do que numa importacao vazia.
    """
    with zipfile.ZipFile(DADOS / "dfp_cia_aberta_2024.zip") as arquivo:
        for nome in arquivo.namelist():
            linhas = [l for l in arquivo.read(nome).split(b"\r\n") if l]
            campos = {l.count(b";") for l in linhas}
            assert len(campos) == 1, f"{nome} tem linhas com contagem de campos diferente"


def test_recorte_em_bytes_pega_so_a_companhia():
    with zipfile.ZipFile(DADOS / "dfp_cia_aberta_2024.zip") as arquivo:
        bruto = arquivo.read("dfp_cia_aberta_DRE_con_2024.csv")

    recorte = _apenas_da_companhia(bruto, WEG)
    assert recorte is not None
    linhas = [l for l in recorte.split(b"\r\n") if l]

    assert linhas[0] == bruto.split(b"\r\n")[0], "o cabecalho tem que sobreviver"
    assert len(linhas) > 1
    for linha in linhas[1:]:
        assert linha.split(b";")[4].lstrip(b"0") == str(WEG).encode()
    # O fixture tem quatro companhias; o recorte tem que ser menor que o todo.
    assert len(linhas) < len([l for l in bruto.split(b"\r\n") if l])


def test_recorte_em_bytes_de_companhia_ausente():
    with zipfile.ZipFile(DADOS / "dfp_cia_aberta_2024.zip") as arquivo:
        bruto = arquivo.read("dfp_cia_aberta_DRE_con_2024.csv")
    assert _apenas_da_companhia(bruto, 999_999) is None


def test_recorte_em_bytes_nao_muda_o_resultado():
    """A otimizacao e invisivel: os numeros tem que ser os mesmos de antes."""
    dfs = importar_cvm(WEG, [2024], cache=DADOS)
    assert dfs.valor("receita_liquida", 2024) == pytest.approx(37_986_941_000.0)
    assert dfs.valor("ativo_total", 2024) == dfs.valor("passivo_total", 2024)


@pytest.mark.parametrize(
    "escala,fator",
    [("MIL", 1_000.0), ("UNIDADE", 1.0), ("mil", 1_000.0), ("MILHAO", 1_000_000.0)],
)
def test_fator_de_escala(escala, fator):
    assert _fator_escala(escala, []) == fator


def test_escala_desconhecida_avisa_em_vez_de_adivinhar():
    avisos: list[str] = []
    assert _fator_escala("BILHAO", avisos) == 1.0
    assert avisos and "ESCALA_MOEDA desconhecida" in avisos[0]


# ---------------------------------------------------------------------------
# Pegadinha 1: ESCALA_MOEDA
# ---------------------------------------------------------------------------


def test_escala_mil_vira_reais(weg):
    """WEG publica em MIL: 37.986.941 no arquivo sao R$ 38 bilhoes."""
    assert weg.valor("receita_liquida", 2024) == pytest.approx(37_986_941_000.0)
    assert weg.valor("lucro_liquido", 2024) == pytest.approx(6_318_763_000.0)


def test_escala_unidade_nao_e_multiplicada(catalogo):
    """Vivara publica em UNIDADE: 2.577.113.417 ja sao R$ 2,6 bilhoes.

    O numero bruto e maior que o da WEG, que vale dez vezes mais. Quem ignora
    ESCALA_MOEDA erra nas duas empresas, e em direcoes opostas.
    """
    vivara = importar_cvm(VIVARA, [2024], cache=DADOS, catalogo=catalogo)
    receita = vivara.valor("receita_liquida", 2024)

    assert receita == pytest.approx(2_577_113_417.0)
    # A armadilha concreta: multiplicar por mil daria R$ 2,6 trilhoes.
    assert receita < 1e12


def test_duas_empresas_de_escalas_diferentes_ficam_comparaveis(weg, catalogo):
    """Depois da conversao, comparar as duas empresas passa a fazer sentido."""
    vivara = importar_cvm(VIVARA, [2024], cache=DADOS, catalogo=catalogo)
    razao = weg.valor("receita_liquida", 2024) / vivara.valor("receita_liquida", 2024)
    assert 10 < razao < 20  # WEG fatura cerca de 15x a Vivara


# ---------------------------------------------------------------------------
# Pegadinha 2: ORDEM_EXERC
# ---------------------------------------------------------------------------


def test_arquivo_anual_contem_o_ano_anterior_como_penultimo():
    """A premissa da pegadinha: o zip de 2024 ja traz 2023 dentro dele."""
    with zipfile.ZipFile(DADOS / "dfp_cia_aberta_2024.zip") as arquivo:
        linhas = arquivo.read("dfp_cia_aberta_DRE_con_2024.csv").decode(ENCODING).splitlines()

    fins = {
        campos[10]
        for linha in linhas[1:]
        if (campos := linha.split(SEPARADOR)) and campos[8] == "PENÚLTIMO"
    }
    assert fins == {"2023-12-31", "2023-03-31"}


def test_anos_nao_duplicam_ao_empilhar_arquivos(weg):
    """Dois zips, dois anos -- nao tres nem quatro."""
    assert weg.anos == [2023, 2024]
    assert len(weg.anos) == len(set(weg.anos))


def test_usa_o_exercicio_ultimo_de_cada_arquivo(catalogo):
    """2023 vem do arquivo de 2023, nao do PENULTIMO do arquivo de 2024.

    Os dois costumam bater, mas so o ULTIMO e o numero como a companhia o
    publicou naquele exercicio; o PENULTIMO ja pode vir reapresentado.
    """
    with zipfile.ZipFile(DADOS / "dfp_cia_aberta_2023.zip") as arquivo:
        linhas = arquivo.read("dfp_cia_aberta_DRE_con_2023.csv").decode(ENCODING).splitlines()

    esperado = next(
        float(campos[13])
        for linha in linhas[1:]
        if (campos := linha.split(SEPARADOR))
        and campos[4].lstrip("0") == str(WEG)
        and campos[11] == "3.01"
        and campos[8] == "ÚLTIMO"
    )

    so_2023 = importar_cvm(WEG, [2023], cache=DADOS, catalogo=catalogo)
    assert so_2023.valor("receita_liquida", 2023) == pytest.approx(esperado * 1_000)


# ---------------------------------------------------------------------------
# Longo -> colunas por ano
# ---------------------------------------------------------------------------


def test_pivo_produz_contas_canonicas_x_anos(weg):
    assert list(weg.valores.columns) == [2023, 2024]
    for chave in ("receita_liquida", "ebit", "lucro_liquido", "ativo_total",
                  "patrimonio_liquido", "caixa_equivalentes"):
        assert weg.tem(chave), f"{chave} deveria ter vindo da CVM"


def test_codigo_cvm_vira_conta_canonica(weg):
    """O reconhecimento vem do plano de contas, nao do texto do rotulo."""
    assert weg.mapeamento["receita_liquida"].startswith("3.01")
    assert weg.mapeamento["patrimonio_liquido"].startswith("2.03")
    assert weg.mapeamento["caixa_equivalentes"].startswith("1.01.01")
    assert weg.mapeamento["ativo_total"].startswith("1 -")


def test_identidade_do_balanco_fecha(weg):
    """Ativo = passivo, exatamente, nos dois anos."""
    for ano in weg.anos:
        assert weg.valor("ativo_total", ano) == weg.valor("passivo_total", ano)

    # A WEG so gera o aviso de arrendamento: ela reporta o passivo de
    # arrendamento fora da subarvore de emprestimos, como 190 das 467
    # companhias de 2024. Qualquer outro aviso seria problema de leitura.
    esperados = ("arrendamento", "capital de giro", "financiamento")
    outros = [a for a in weg.avisos if not any(e in a for e in esperados)]
    assert not outros, f"aviso inesperado na WEG: {outros}"


def test_custos_viram_magnitude_positiva(weg):
    """A CVM publica CPV negativo; o motor subtrai a conta, entao guarda o modulo."""
    assert weg.valor("custo_produtos_vendidos", 2024) > 0
    assert weg.valor("impostos", 2024) > 0
    # E o resultado bruto continua fechando com receita - CPV.
    assert weg.valor("lucro_bruto", 2024) == pytest.approx(
        weg.valor("receita_liquida", 2024) - weg.valor("custo_produtos_vendidos", 2024)
    )


def test_linha_da_dfc_nao_invade_conta_do_balanco(weg):
    """A DFC da WEG tem uma linha chamada so "Imobilizado" -- que e capex.

    Sem amarrar o reconhecimento a demonstracao de origem, esse fluxo de caixa
    sobrescreveria o saldo de imobilizado do balanco.
    """
    assert weg.mapeamento["imobilizado"].startswith("1.02.03")
    assert weg.valor("imobilizado", 2024) == pytest.approx(9_933_659_000.0)


def test_nada_e_descartado_em_silencio(weg):
    """Toda linha publicada esta na arvore, mapeada ou nao.

    Uma conta filha nao esta "nao reconhecida": ela e a abertura de uma conta
    que o app entende, e o lugar dela e a arvore. O que sobra na lista de nao
    reconhecidas e so o que nao tem pai nem conta canonica.
    """
    publicadas = set(weg.detalhe["codigo"])
    assert len(publicadas) > 200

    # As analiticas continuam presentes, agora com hierarquia.
    assert "1.01.06" in publicadas  # tributos a recuperar
    assert "2.03.02" in publicadas  # reservas de capital
    assert "3.04.01" in publicadas  # despesas com vendas

    # Comparar por contagem de linhas, e nao por rotulo: "Outros" e "Outras
    # Obrigacoes" aparecem em varios pontos do plano e nao sao a mesma conta.
    assert len(weg.arvore()) == len(publicadas)


def test_a_arvore_tem_a_hierarquia_do_plano(weg):
    """O recuo vem do nivel do codigo, e a ordem e a da demonstracao."""
    bp = weg.arvore("bp")
    rotulos = list(bp.index)

    assert rotulos[0].strip().startswith("Ativo Total")
    assert not rotulos[0].startswith(" "), "conta de primeiro nivel nao tem recuo"

    circulante = next(i for i, r in enumerate(rotulos) if "Ativo Circulante" in r)
    caixa = next(i for i, r in enumerate(rotulos) if "Caixa e Equivalentes" in r)
    assert circulante < caixa, "a filha tem que vir depois da mae"
    assert len(rotulos[caixa]) - len(rotulos[caixa].lstrip()) > (
        len(rotulos[circulante]) - len(rotulos[circulante].lstrip())
    ), "a filha tem que ter mais recuo que a mae"


def test_a_arvore_ordena_por_hierarquia_e_nao_por_texto(weg):
    """Em ordem alfabetica 1.01.10 viria antes de 1.01.02."""
    from valuation.importacao.cvm import _ordem_do_codigo

    assert _ordem_do_codigo("1.01.02") < _ordem_do_codigo("1.01.10")
    assert _ordem_do_codigo("1.01") < _ordem_do_codigo("1.01.01")
    assert _ordem_do_codigo("1") < _ordem_do_codigo("2")


def test_conta_publicada_com_grafias_diferentes_nao_duplica(weg):
    """A mesma conta muda de caixa entre anos: "Receita Liquida" e "Receita liquida"."""
    codigos = list(weg.detalhe["codigo"])
    assert len(codigos) == len(set(codigos)), "codigo repetido na arvore"


def test_a_arvore_soma_de_baixo_para_cima(weg):
    """A quebra explica o total: 1.01 e a soma das filhas 1.01.x."""
    valores = dict(zip(weg.detalhe["codigo"], weg.detalhe[2024]))
    filhas = [
        v
        for c, v in valores.items()
        if c.startswith("1.01.") and c.count(".") == 2 and np.isfinite(v)
    ]
    assert filhas
    assert sum(filhas) == pytest.approx(valores["1.01"], rel=0.005)


def test_composicao_soma_o_total(weg):
    """A quebra tem que somar o pai: e o que a torna leitura e nao ilustracao."""
    comp = weg.composicao("1.01", 2024)
    assert not comp.empty
    assert comp["Valor"].sum() == pytest.approx(weg.valor("ativo_circulante", 2024))
    assert comp["% do total"].sum() == pytest.approx(1.0)


def test_composicao_so_pega_filhas_diretas(weg):
    """1.01.01.01 compoe 1.01.01, nao 1.01 -- somaria em dobro."""
    codigos = list(weg.composicao("1.01", 2024).index)
    assert all(c.count(".") == 2 for c in codigos)
    assert "1.01.01" in codigos
    assert "1.01.01.01" not in codigos


def test_composicao_responde_de_que_e_feita_a_divida(weg):
    """Emprestimo, debenture ou arrendamento -- o total nao distingue."""
    comp = weg.composicao("2.01.04", 2024)
    assert set(comp.index) >= {"2.01.04.01", "2.01.04.02", "2.01.04.03"}

    # A subarvore publicada explica *parte* da divida de curto prazo. O resto e
    # arrendamento que a companhia reportou fora dela e que o leitor devolve --
    # ver _somar_arrendamento_fora_da_divida. Antes desta correcao os dois
    # numeros batiam, e batiam errado.
    subarvore = comp["Valor"].sum()
    canonica = weg.valor("divida_curto_prazo", 2024)
    assert canonica > subarvore
    assert canonica - subarvore == pytest.approx(
        weg.valor("arrendamento_curto_prazo", 2024) - comp.loc["2.01.04.03", "Valor"]
    )


def test_composicao_de_conta_sem_filhas(weg):
    assert weg.composicao("1.01.05", 2024).empty
    assert weg.composicao("9.99", 2024).empty


def test_indicadores_de_liquidez(weg):
    analise = analisar(weg)
    for indicador in (
        "Liquidez corrente",
        "Liquidez seca",
        "Liquidez imediata",
        "FCO / Passivo circulante",
        "Caixa / Divida de curto prazo",
    ):
        assert indicador in analise.indicadores.index, indicador

    corrente = analise.ultimo("Liquidez corrente")
    seca = analise.ultimo("Liquidez seca")
    assert seca < corrente, "a seca tira o estoque e tem que ser menor"
    assert corrente == pytest.approx(
        weg.valor("ativo_circulante", 2024) / weg.valor("passivo_circulante", 2024)
    )


def test_a_arvore_acompanha_a_escala(weg):
    milhoes = weg.escalar(1_000_000, "R$ milhões")
    original = dict(zip(weg.detalhe["codigo"], weg.detalhe[2024]))
    escalada = dict(zip(milhoes.detalhe["codigo"], milhoes.detalhe[2024]))
    assert escalada["1"] == pytest.approx(original["1"] / 1_000_000)
    # Codigo e nivel nao sao valores e nao podem ter sido divididos.
    assert list(milhoes.detalhe["nivel"]) == list(weg.detalhe["nivel"])


# ---------------------------------------------------------------------------
# Classificacao detalhada: DRE, BP e DFC abertos
# ---------------------------------------------------------------------------


def test_reconhece_a_demonstracao_inteira(weg):
    """O objetivo nao e achar as contas do modelo, e ler a DF publicada."""
    assert len(weg.valores.index) >= 60, "a cobertura do plano de contas regrediu"
    for demonstracao, minimo in (("dre", 15), ("bp", 25), ("dfc", 8)):
        assert len(weg.tabela(demonstracao).index) >= minimo, demonstracao


def test_capex_soma_imobilizado_e_intangivel(weg):
    """A WEG separa em duas linhas; guardar so uma subestima o investimento.

    6.02.02 Imobilizado (1.780.663) + 6.02.03 Intangivel (69.659), em milhares.
    """
    assert weg.valor("capex", 2024) == pytest.approx(1_850_322_000.0)
    assert " + " in weg.mapeamento["capex"]
    assert "6.02.02" in weg.mapeamento["capex"]
    assert "6.02.03" in weg.mapeamento["capex"]


def test_capex_nao_engole_a_venda_de_imobilizado(weg):
    """6.02.04 e recebimento pela venda de imobilizado -- entrada, nao capex."""
    assert "6.02.04" not in weg.mapeamento["capex"]
    # Continua publicada e visivel na arvore, so nao somada ao capex.
    assert "6.02.04" in set(weg.detalhe["codigo"])


def test_juros_pagos_somados_das_duas_secoes(weg):
    """A WEG paga juros dentro do operacional e do financiamento."""
    assert weg.tem("juros_pagos")
    assert weg.valor("juros_pagos", 2024) == pytest.approx(160_301_000.0)
    assert weg.mapeamento["juros_pagos"].count(" + ") >= 1


def test_juros_pagos_nao_confunde_jcp_com_juros_de_divida(weg):
    """JCP e remuneracao ao acionista; entra em dividendos, nao no custo da divida."""
    assert "Capital Próprio" not in weg.mapeamento["juros_pagos"]
    assert "Capital Próprio" in weg.mapeamento["dividendos_pagos"]


def test_dividendos_pagos_separado_dos_recebidos(weg):
    assert weg.valor("dividendos_pagos", 2024) == pytest.approx(2_934_611_000.0)
    assert "recebid" not in weg.mapeamento["dividendos_pagos"].lower()


def test_investimento_em_capital_de_giro_vem_da_dfc(weg):
    """6.01.02 e o giro medido pelo caixa, **sem** o que nao e giro.

    A WEG lanca juros e impostos pagos dentro de 6.01.02. Lido cru, o giro dela
    consome R$ 2.310 mi; tirados os pagamentos, consome R$ 774 mi. O primeiro
    numero nao e investimento em giro -- e giro mais desembolso de juro e de
    imposto --, e ele alimentava premissa de projecao.
    """
    publicado = -2_310_041_000.0
    reclassificado = weg.valor("pagamentos_reclassificados_do_giro", 2024)
    assert reclassificado == pytest.approx(1_535_663_000.0)

    giro = weg.valor("variacao_capital_giro", 2024)
    assert giro == pytest.approx(publicado + reclassificado)
    # Negativo: o giro consumiu caixa no ano, so que menos do que parecia.
    assert giro < 0


def test_fco_se_abre_em_geracao_e_giro(weg):
    """Geracao + giro - pagamentos reclassificados = caixa operacional.

    A decomposicao ganhou um termo: o que foi tirado do giro continua dentro do
    FCO, so que abaixo dele. Sem o termo, a soma nao fecha -- e nao fechar aqui
    seria sinal de que a reclassificacao vazou caixa para fora do operacional.
    """
    soma = (
        weg.valor("caixa_das_operacoes", 2024)
        + weg.valor("variacao_capital_giro", 2024)
        - weg.valor("pagamentos_reclassificados_do_giro", 2024)
    )
    # "Outros" (6.01.03) fecha a diferenca; a geracao explica a maior parte.
    assert soma == pytest.approx(weg.valor("fluxo_operacional", 2024), rel=0.01)


def test_a_dfc_fecha(weg):
    """Operacional + investimento + financiamento + cambio = variacao do caixa."""
    for ano in weg.anos:
        soma = (
            weg.valor("fluxo_operacional", ano)
            + weg.valor("fluxo_investimento", ano)
            + weg.valor("fluxo_financiamento", ano)
            + weg.valor("variacao_cambial_caixa", ano)
        )
        assert soma == pytest.approx(weg.valor("variacao_caixa", ano), rel=1e-6)


def test_lucro_se_reparte_entre_controladores_e_minoritarios(weg):
    assert weg.valor("lucro_controladores", 2024) + weg.valor(
        "lucro_nao_controladores", 2024
    ) == pytest.approx(weg.valor("lucro_liquido", 2024))


def test_divida_aberta_em_debentures_e_arrendamento(catalogo):
    """Arrendamento e debenture sao filhas de emprestimos, nao parcelas a somar."""
    dfs = importar_cvm(SAO_MARTINHO, [2024], cache=DADOS, catalogo=catalogo)
    for filha, mae in (
        ("arrendamento_curto_prazo", "divida_curto_prazo"),
        ("debentures_curto_prazo", "divida_curto_prazo"),
        ("arrendamento_longo_prazo", "divida_longo_prazo"),
        ("debentures_longo_prazo", "divida_longo_prazo"),
    ):
        if dfs.tem(filha) and dfs.tem(mae):
            assert dfs.valor(filha, 2024) <= dfs.valor(mae, 2024) * 1.001, filha


def test_direito_de_uso_esta_dentro_do_imobilizado(weg):
    assert weg.valor("direito_uso_arrendamento", 2024) <= weg.valor("imobilizado", 2024)


def test_goodwill_esta_dentro_do_intangivel(weg):
    assert weg.valor("goodwill", 2024) <= weg.valor("intangivel", 2024)


# ---------------------------------------------------------------------------
# Conferencia conta a conta contra a linha publicada
# ---------------------------------------------------------------------------

FATOR_ESCALA = {"MIL": 1_000.0, "UNIDADE": 1.0, "MILHAO": 1_000_000.0}


def _rotulo_publicado(dfs, codigo: str) -> str:
    """Rotulo de uma linha publicada, no formato que a planilha usa."""
    linha = dfs.detalhe[dfs.detalhe["codigo"] == codigo].iloc[0]
    return f"{linha['codigo']} - {linha['rotulo']}"


def _linhas_publicadas(codigo_cvm: int, escopo: str, ano: int) -> dict:
    """Todas as linhas da companhia lidas do zip sem passar pelo importador."""
    publicadas = {}
    with zipfile.ZipFile(DADOS / f"dfp_cia_aberta_{ano}.zip") as arquivo:
        for grupo in ("DRE", "BPA", "BPP", "DFC_MI", "DFC_MD"):
            nome = f"dfp_cia_aberta_{grupo}_{escopo}_{ano}.csv"
            if nome not in arquivo.namelist():
                continue
            for linha in arquivo.read(nome).decode(ENCODING).splitlines()[1:]:
                campos = linha.split(SEPARADOR)
                if campos[4].lstrip("0") != str(codigo_cvm):
                    continue
                # BPA e BPP nao tem DT_INI_EXERC: as colunas do fim andam uma casa.
                deslocamento = 0 if len(campos) == 15 else -1
                if campos[8] != "ÚLTIMO":
                    continue
                escala = FATOR_ESCALA[campos[7].strip().upper()]
                publicadas[campos[11 + deslocamento].strip()] = (
                    float(campos[13 + deslocamento]) * escala
                )
    return publicadas


@pytest.mark.parametrize(
    "codigo,escopo", [(WEG, "con"), (VIVARA, "con"), (SAO_MARTINHO, "con"), (ELEKTRO, "ind")]
)
def test_cada_conta_bate_com_a_linha_publicada(codigo, escopo, catalogo):
    """A verificacao mais forte que existe: conta a conta contra o arquivo.

    Nao confia em nenhuma etapa do importador. Volta ao CSV bruto, acha a linha
    pelo codigo que o app registrou no mapeamento, aplica escala e sinal a mao e
    compara. Escala errada, sinal trocado, conta trocada ou agregacao que somou
    o que nao devia -- tudo aparece aqui como diferenca.
    """
    dfs = importar_cvm(codigo, [2024], cache=DADOS, catalogo=catalogo)
    publicadas = _linhas_publicadas(codigo, escopo, 2024)
    assert publicadas, "o fixture nao tem linhas desta companhia"

    conferidas = 0
    for chave in dfs.valores.index:
        obtido = dfs.valor(chave, 2024)
        if not np.isfinite(obtido):
            continue
        origem = dfs.mapeamento.get(chave)
        if not origem:  # conta derivada, nao publicada
            continue
        # Contas que o app **reclassifica de proposito** nao batem com a linha
        # publicada, e e esse o ponto delas. Cada uma tem teste proprio.
        if "(" in origem or chave in _RECLASSIFICADAS:
            continue

        codigos = [parte.split(" - ")[0].strip() for parte in origem.split(" + ")]
        if not all(c in publicadas for c in codigos):
            continue

        esperado = sum(publicadas[c] for c in codigos)
        if chave == "impostos":
            # ``impostos`` guarda **despesa positiva**, e o sinal sai da
            # identidade da companhia e nao da magnitude: quem publicou credito
            # (3.08 positivo) fica negativo aqui. Sao 118 das 467 companhias de
            # 2024. Ver ``_corrigir_sinal_dos_impostos``.
            esperado = -esperado
        elif POR_CHAVE[chave].sinal_invertido:
            esperado = abs(esperado)
        conferidas += 1
        assert obtido == pytest.approx(esperado, rel=1e-9), (
            f"{chave} veio de {origem}: app={obtido}, arquivo={esperado}"
        )

    assert conferidas >= 25, f"conferiu poucas contas ({conferidas})"


def test_hierarquia_do_plano_de_contas_soma(catalogo):
    """Pai = soma dos filhos, no proprio arquivo da CVM.

    Valida de uma vez o arquivo e a leitura da escala: se a escala fosse
    aplicada de forma inconsistente entre linhas, a soma deixaria de fechar.
    """
    publicadas = _linhas_publicadas(WEG, "con", 2024)

    filhos: dict[str, list[str]] = {}
    for codigo in publicadas:
        if "." in codigo:
            filhos.setdefault(codigo.rsplit(".", 1)[0], []).append(codigo)

    verificados = 0
    for pai, lista in filhos.items():
        # 6.05 tem saldo inicial e final como "filhos", e e a diferenca entre
        # eles; 3.99 e lucro por acao, valor unitario. Nenhum dos dois soma.
        if pai not in publicadas or len(lista) < 2 or pai[:4] in {"6.05", "3.99"}:
            continue
        soma = sum(publicadas[c] for c in lista)
        # Companhia que informa o total e deixa as filhas zeradas nao detalhou a
        # conta; ausencia de abertura nao e inconsistencia.
        if soma == 0 and publicadas[pai] != 0:
            continue
        verificados += 1
        assert soma == pytest.approx(publicadas[pai], rel=0.005), (
            f"{pai}: pai={publicadas[pai]}, filhos somam {soma}"
        )

    assert verificados >= 30


# ---------------------------------------------------------------------------
# Planos de contas diferentes
# ---------------------------------------------------------------------------


def test_banco_e_reconhecido_como_outro_plano_de_contas(catalogo):
    """No plano financeiro os mesmos codigos sao outras contas.

    3.06 e "Resultado Financeiro" na industria e "IR e CSLL" no banco. Ler pelo
    codigo poria o imposto na linha do resultado financeiro sem avisar.
    """
    dfs = importar_cvm(BANCO_BRASIL, [2024], cache=DADOS, catalogo=catalogo)
    assert any("instituicao financeira" in aviso for aviso in dfs.avisos)

    # A linha 3.06 e o desvio que motivou tudo: no plano industrial e resultado
    # financeiro, aqui e o imposto. Ela tem que ter chegado em "impostos".
    donos_do_306 = [c for c, o in dfs.mapeamento.items() if o.startswith("3.06 ")]
    assert donos_do_306 == ["impostos"], donos_do_306


def test_banco_reconhece_pelo_plano_proprio(catalogo):
    """No plano financeiro 2.07 e patrimonio liquido, e 2.03 sao provisoes."""
    dfs = importar_cvm(BANCO_BRASIL, [2024], cache=DADOS, catalogo=catalogo)

    for chave in ("ativo_total", "passivo_total", "patrimonio_liquido",
                  "caixa_equivalentes", "receita_liquida", "lucro_liquido"):
        assert dfs.tem(chave), chave

    assert dfs.mapeamento["patrimonio_liquido"].startswith("2.07")
    assert dfs.mapeamento["caixa_equivalentes"].startswith("1.01")
    # A identidade do balanco tem que fechar tambem no plano financeiro.
    assert dfs.valor("ativo_total", 2024) == pytest.approx(
        dfs.valor("passivo_total", 2024)
    )
    assert dfs.valor("patrimonio_liquido", 2024) < dfs.valor("ativo_total", 2024)


def test_banco_nao_ganha_ebit_inventado(catalogo):
    """Banco nao tem EBIT nem capital de giro operacional; fingir seria pior."""
    dfs = importar_cvm(BANCO_BRASIL, [2024], cache=DADOS, catalogo=catalogo)
    assert not dfs.tem("ebit") or dfs.mapeamento.get("ebit") is None
    # E o aviso de que o modelo nao se aplica continua.
    assert any("instituicao financeira" in a for a in dfs.avisos)


def test_o_mapa_financeiro_nao_afeta_o_plano_industrial(weg):
    """3.06 e resultado financeiro na industria, imposto no banco."""
    assert weg.mapeamento["resultado_financeiro"].startswith("3.06")
    assert weg.mapeamento["impostos"].startswith("3.08")


def test_plano_industrial_nao_dispara_o_aviso(weg):
    assert not any("instituicao financeira" in aviso for aviso in weg.avisos)


def test_deteccao_do_plano_e_por_companhia():
    from valuation.importacao.cvm import (
        PLANO_FINANCEIRO,
        PLANO_INDUSTRIAL,
        LinhaCVM,
        detectar_plano,
    )

    def linha(codigo, descricao):
        return LinhaCVM(codigo, descricao, 1.0, 2024, "dre", "MIL", "con")

    assert detectar_plano([linha("3.01", "Receita de Venda de Bens e/ou Serviços")]) == (
        PLANO_INDUSTRIAL
    )
    assert detectar_plano([linha("3.01", "Receitas de Intermediação Financeira")]) == (
        PLANO_FINANCEIRO
    )
    assert detectar_plano([linha("3.01", "Receitas das Atividades Seguradoras")]) == (
        PLANO_FINANCEIRO
    )


# ---------------------------------------------------------------------------
# Casos reais que quebram a suposicao ingenua
# ---------------------------------------------------------------------------


def test_exercicio_social_que_nao_fecha_em_dezembro(catalogo):
    """Sao Martinho fecha em 31/03. O ano e o do encerramento, nao o do arquivo."""
    dfs = importar_cvm(SAO_MARTINHO, [2023, 2024], cache=DADOS, catalogo=catalogo)
    assert dfs.anos == [2023, 2024]
    assert dfs.valor("receita_liquida", 2024) == pytest.approx(6_891_738_000.0)


def test_companhia_sem_consolidado_cai_para_o_individual(catalogo):
    """Elektro Redes so publica individual -- e o app precisa dizer isso."""
    dfs = importar_cvm(ELEKTRO, [2024], cache=DADOS, catalogo=catalogo)
    assert dfs.valor("receita_liquida", 2024) == pytest.approx(9_328_000_000.0)
    assert any("individual" in aviso for aviso in dfs.avisos)


def test_escopo_e_o_mesmo_em_todas_as_demonstracoes():
    """DRE do grupo com DFC da empresa isolada seriam duas entidades na tabela.

    O escopo e resolvido uma vez por companhia, antes de ler qualquer
    demonstracao. Nos arquivos de 2024, nenhuma das 467 companhias com
    consolidado deixa de publicar alguma demonstracao nesse escopo, entao
    travar aqui nao custa dado.
    """
    from valuation.importacao.cvm import escopo_da_companhia

    zip_2024 = DADOS / "dfp_cia_aberta_2024.zip"
    assert escopo_da_companhia(zip_2024, 2024, WEG) == "con"
    assert escopo_da_companhia(zip_2024, 2024, ELEKTRO) == "ind"
    assert escopo_da_companhia(zip_2024, 2024, 999_999) is None


def test_companhia_com_consolidado_nao_mistura_com_individual(weg):
    """A WEG publica nos dois escopos; nenhuma linha pode vir do individual."""
    assert not any("individual" in aviso for aviso in weg.avisos)
    # O ativo total consolidado da WEG em 2024 e maior que o individual.
    assert weg.valor("ativo_total", 2024) == pytest.approx(41_489_701_000.0)


def test_companhia_inexistente_explica_o_que_houve(catalogo):
    with pytest.raises(ErroCVM, match="nao tem DFP desta companhia"):
        importar_cvm(999_999, [2024], cache=DADOS, catalogo=catalogo)


def test_ano_sem_arquivo_no_cache_nao_inventa_numero():
    with pytest.raises(ErroCVM):
        importar_cvm(WEG, [1998], cache=DADOS)


def test_lista_de_anos_vazia_e_erro():
    with pytest.raises(ErroCVM, match="ao menos um ano"):
        importar_cvm(WEG, [], cache=DADOS)


# ---------------------------------------------------------------------------
# Cadastro e busca
# ---------------------------------------------------------------------------


def test_cadastro_carrega_companhias(catalogo):
    codigos = {c.codigo_cvm for c in catalogo}
    assert codigos == {WEG, VIVARA, SAO_MARTINHO, ELEKTRO, BANCO_BRASIL}
    assert all(c.ativa for c in catalogo)


def test_busca_por_nome(catalogo):
    achados = buscar_companhias("weg", catalogo)
    assert [c.codigo_cvm for c in achados] == [WEG]


def test_busca_ignora_acento_e_caixa(catalogo):
    assert buscar_companhias("sao martinho", catalogo)[0].codigo_cvm == SAO_MARTINHO
    assert buscar_companhias("SÃO MARTINHO", catalogo)[0].codigo_cvm == SAO_MARTINHO


@pytest.mark.parametrize("termo", ["84.429.695/0001-11", "84429695000111"])
def test_busca_por_cnpj_com_ou_sem_pontuacao(catalogo, termo):
    assert buscar_companhias(termo, catalogo)[0].codigo_cvm == WEG


def test_busca_vazia_nao_devolve_o_catalogo_inteiro(catalogo):
    assert buscar_companhias("", catalogo) == []
    assert buscar_companhias("   ", catalogo) == []


def test_busca_sem_resultado(catalogo):
    assert buscar_companhias("empresa que nao existe", catalogo) == []


def test_quem_comeca_com_o_termo_vem_antes(catalogo):
    """Buscar "vivara" tem que trazer a Vivara, nao quem so a menciona."""
    achados = buscar_companhias("vivara", catalogo)
    assert achados[0].codigo_cvm == VIVARA


def test_cadastro_traz_cnpj_e_setor(catalogo):
    weg = next(c for c in catalogo if c.codigo_cvm == WEG)
    assert weg.cnpj == "84.429.695/0001-11"
    assert weg.setor
    assert "WEG" in str(weg)


# ---------------------------------------------------------------------------
# Cache
# ---------------------------------------------------------------------------


def test_arquivo_em_cache_nao_e_rebaixado(monkeypatch):
    """Com o zip em disco, nenhuma conexao e aberta."""
    def explodir(*args, **kwargs):
        raise AssertionError("tentou baixar um arquivo que ja estava em cache")

    monkeypatch.setattr("urllib.request.urlopen", explodir)
    caminho = baixar_dfp(2024, cache=DADOS)
    assert caminho.exists()
    assert caminho.name == "dfp_cia_aberta_2024.zip"


def test_download_grava_e_reaproveita(monkeypatch, tmp_path):
    chamadas: list[str] = []

    class RespostaFalsa:
        def __init__(self, dado):
            self._dado = dado

        def read(self):
            return self._dado

        def __enter__(self):
            return self

        def __exit__(self, *args):
            return False

    def falsa(url, timeout=None):
        chamadas.append(url)
        return RespostaFalsa(b"conteudo-zip")

    monkeypatch.setattr("urllib.request.urlopen", falsa)

    primeiro = baixar_dfp(2019, cache=tmp_path)
    assert primeiro.read_bytes() == b"conteudo-zip"
    assert len(chamadas) == 1

    baixar_dfp(2019, cache=tmp_path)
    assert len(chamadas) == 1, "o segundo pedido deveria sair do cache"


def test_download_interrompido_nao_deixa_arquivo_truncado(monkeypatch, tmp_path):
    """Meio arquivo em disco seria lido como valido na proxima execucao."""
    import urllib.error

    def falhar(url, timeout=None):
        raise urllib.error.URLError("conexao caiu")

    monkeypatch.setattr("urllib.request.urlopen", falhar)

    with pytest.raises(ErroCVM, match="tentativas"):
        baixar_dfp(2019, cache=tmp_path)

    assert list(tmp_path.iterdir()) == []


def test_conexao_cortada_no_meio_vira_mensagem_e_nao_crash(monkeypatch, tmp_path):
    """``IncompleteRead`` nao e ``URLError`` -- e o portal corta zips de 13 MB.

    Sem tratar esta excecao especifica, o corte no meio do download subia como
    erro nao tratado ate a tela do usuario. Aconteceu de verdade baixando o
    arquivo de 2024.
    """
    import http.client

    def cortar(url, timeout=None):
        raise http.client.IncompleteRead(b"come" * 10, 2_952_985)

    monkeypatch.setattr("urllib.request.urlopen", cortar)

    with pytest.raises(ErroCVM, match="corta a conexao"):
        baixar_dfp(2024, cache=tmp_path)
    assert list(tmp_path.iterdir()) == []


def test_download_tenta_de_novo_antes_de_desistir(monkeypatch, tmp_path):
    """Falha transitoria nao pode custar um clique ao usuario."""
    import http.client

    chamadas = []

    class Resposta:
        def read(self):
            return b"zip-de-verdade"

        def __enter__(self):
            return self

        def __exit__(self, *args):
            return False

    def instavel(url, timeout=None):
        chamadas.append(url)
        if len(chamadas) == 1:
            raise http.client.IncompleteRead(b"", 100)
        return Resposta()

    monkeypatch.setattr("urllib.request.urlopen", instavel)

    caminho = baixar_dfp(2019, cache=tmp_path)
    assert caminho.read_bytes() == b"zip-de-verdade"
    assert len(chamadas) == 2, "deveria ter tentado duas vezes"
    assert not list(tmp_path.glob("*.parcial"))


def test_tamanho_do_cache(tmp_path):
    from valuation.importacao.cvm import tamanho_do_cache

    assert tamanho_do_cache(tmp_path) == 0
    (tmp_path / "dfp_cia_aberta_2019.zip").write_bytes(b"x" * 1000)
    (tmp_path / "cad_cia_aberta.csv").write_bytes(b"y" * 500)
    assert tamanho_do_cache(tmp_path) == 1500


def test_limpar_cache_preserva_o_cadastro(tmp_path):
    """O cadastro e pequeno e usado em toda busca; limpar nao deve custa-lo."""
    from valuation.importacao.cvm import limpar_cache

    (tmp_path / "dfp_cia_aberta_2019.zip").write_bytes(b"x")
    (tmp_path / "dfp_cia_aberta_2020.zip").write_bytes(b"x")
    (tmp_path / "cad_cia_aberta.csv").write_bytes(b"y")

    assert limpar_cache(tmp_path) == 2
    assert (tmp_path / "cad_cia_aberta.csv").exists()
    assert not (tmp_path / "dfp_cia_aberta_2019.zip").exists()

    assert limpar_cache(tmp_path, manter_cadastro=False) == 1
    assert not (tmp_path / "cad_cia_aberta.csv").exists()


def test_limpar_cache_inexistente_nao_quebra(tmp_path):
    from valuation.importacao.cvm import limpar_cache, tamanho_do_cache

    ausente = tmp_path / "nao_existe"
    assert tamanho_do_cache(ausente) == 0
    assert limpar_cache(ausente) == 0


def test_zip_corrompido_sugere_limpar_o_cache(tmp_path):
    (tmp_path / "dfp_cia_aberta_2024.zip").write_bytes(b"isto nao e um zip")
    with pytest.raises(ErroCVM, match="cache"):
        importar_cvm(WEG, [2024], cache=tmp_path)


# ---------------------------------------------------------------------------
# Integracao com o resto do motor
# ---------------------------------------------------------------------------


def test_demonstracoes_da_cvm_alimentam_a_analise_historica(weg):
    analise = analisar(weg)
    assert not analise.indicadores.empty

    sugestao = sugerir_premissas(analise, horizonte=5)
    assert sugestao.operacionais.receita_base == pytest.approx(37_986_941_000.0)
    assert sugestao.operacionais.ano_base == 2024
    # Divida bruta = curto + longo prazo, direto do balanco da CVM.
    assert sugestao.ponte.divida_bruta == pytest.approx(
        weg.valor("divida_curto_prazo", 2024) + weg.valor("divida_longo_prazo", 2024)
    )


def test_custo_da_divida_vem_do_juro_pago_e_nao_da_despesa_financeira(weg):
    """A linha 3.06.02 da CVM nao e juro de divida.

    Ela junta variacao cambial e monetaria de todo o passivo. Na WEG de 2024 da
    R$ 1,72 bi sobre R$ 3,6 bi de divida -- 48% ao ano --, enquanto o juro que
    saiu do caixa foi R$ 160 mi. Usar a primeira inflava o WACC e derrubava o
    valor sem avisar; medido na base de 2024, em 28% das companhias.
    """
    analise = analisar(weg)
    pela_dre = analise.mediana("Custo da divida efetivo")
    pelo_caixa = analise.mediana("Custo da divida pelo caixa")

    assert pela_dre > 0.30, "o problema que motivou a correcao sumiu do fixture"
    assert 0 < pelo_caixa < 0.15

    sugestao = sugerir_premissas(analise, horizonte=5)
    assert sugestao.custo_capital.custo_divida_brl == pytest.approx(pelo_caixa)
    assert "juros pagos" in sugestao.justificativas["custo_capital"]


def test_kd_implausivel_nao_e_propagado(weg):
    """Sem juros pagos, um Kd absurdo vira Kd sintetico e um alerta."""
    from dataclasses import replace

    sem_juros = type(weg)(
        **{
            **weg.__dict__,
            "valores": weg.valores.drop(index=["juros_pagos"]),
        }
    )
    sugestao = sugerir_premissas(analisar(sem_juros), horizonte=5)
    assert sugestao.custo_capital.custo_divida_brl is None
    assert any("custo da divida" in alerta.lower() for alerta in sugestao.alertas)


def test_acoes_em_circulacao_vem_do_arquivo_de_capital(weg):
    """Emitidas menos tesouraria: 4.197.317.998 - 1.780.620 na WEG de 2024."""
    assert weg.valor("acoes_em_circulacao", 2024) == pytest.approx(4_195_537_378.0)


def test_acoes_acompanham_a_escala_dos_valores(weg):
    """Equity dividido por acoes tem que dar o mesmo preco em qualquer unidade."""
    em_reais = weg.valor("patrimonio_liquido", 2024) / weg.valor(
        "acoes_em_circulacao", 2024
    )
    milhoes = weg.escalar(1_000_000, "R$ milhões")
    em_milhoes = milhoes.valor("patrimonio_liquido", 2024) / milhoes.valor(
        "acoes_em_circulacao", 2024
    )
    assert em_reais == pytest.approx(em_milhoes)
    assert em_reais == pytest.approx(5.51, abs=0.01)


def test_acoes_nao_disparam_aviso_de_escala(weg):
    """A quantidade nao e valor monetario e nao entra na checagem de escala."""
    assert not any("escala mudou" in aviso for aviso in weg.avisos)


def test_acoes_preenchem_a_ponte_sem_o_usuario_digitar(weg):
    sugestao = sugerir_premissas(analisar(weg), horizonte=5)
    assert sugestao.ponte.acoes_em_circulacao == pytest.approx(4_195_537_378.0)
    assert "Acoes em circulacao" in sugestao.justificativas["ponte"]


def test_setor_viaja_na_fonte_para_montar_o_peer_group(catalogo):
    dfs = importar_cvm(WEG, [2024], cache=DADOS, catalogo=catalogo)
    assert dfs.fonte["setor"]


def test_indicadores_de_caixa_aparecem_quando_ha_dfc(weg):
    analise = analisar(weg)
    for indicador in (
        "Conversao de caixa (FCO / EBITDA)",
        "Capex / FCO",
        "Custo da divida pelo caixa",
        "Investimento em giro (DFC) / Receita",
        "Payout (dividendos / lucro)",
    ):
        assert indicador in analise.indicadores.index, indicador


def test_valores_saem_em_reais_e_a_escala_e_do_app(weg):
    assert weg.unidade == "R$"
    assert weg.moeda == "BRL"

    milhoes = weg.escalar(1_000_000, "R$ milhões")
    assert milhoes.valor("receita_liquida", 2024) == pytest.approx(37_986.941)
    assert milhoes.unidade == "R$ milhões"


def test_origem_identifica_a_fonte(weg):
    assert "CVM" in weg.origem and "2023" in weg.origem and "2024" in weg.origem


def test_fonte_guarda_como_rebuscar(weg):
    """``origem`` e frase para ler; ``fonte`` e o suficiente para repetir a busca."""
    assert weg.fonte == {"tipo": "cvm", "codigo_cvm": WEG, "anos": [2023, 2024]}

    # Refazer a busca a partir da fonte tem que devolver os mesmos numeros.
    de_novo = importar_cvm(
        weg.fonte["codigo_cvm"], weg.fonte["anos"], cache=DADOS
    )
    assert de_novo.valores.equals(weg.valores)


def test_fonte_sobrevive_a_troca_de_unidade(weg):
    """``escalar`` cria outro objeto; perder a fonte ali quebraria o botao de atualizar."""
    assert weg.escalar(1_000_000, "R$ milhões").fonte == weg.fonte


def test_fonte_sobrevive_a_correcao_manual(tmp_path):
    destino = tmp_path / "weg.xlsx"
    dfs = importar_cvm(WEG, [2024], cache=DADOS, planilha=destino)
    alvo = _rotulo_publicado(dfs, "1.02.01")
    corrigido = aplicar_mapeamento_manual(dfs, destino, {alvo: "aplicacoes_financeiras"})
    assert corrigido.fonte == dfs.fonte


# ---------------------------------------------------------------------------
# A planilha de conferencia, que e o que liga a CVM a tela de correcao manual
# ---------------------------------------------------------------------------


def test_planilha_gerada_tem_uma_aba_por_demonstracao(tmp_path):
    destino = tmp_path / "weg.xlsx"
    importar_cvm(WEG, [2023, 2024], cache=DADOS, planilha=destino)

    abas = carregar_abas(destino)
    assert set(abas) == {"DRE", "Balanço", "DFC"}

    dre = abas["DRE"]
    assert list(dre.iloc[0]) == ["Código", "Conta", 2023, 2024]
    assert (dre.iloc[:, 1] == "3.01 - Receita de Venda de Bens e/ou Serviços").any()


def test_planilha_permite_corrigir_um_mapeamento_a_mao(tmp_path):
    """E o caminho que a tela de conferencia usa quando o usuario remapeia.

    Sem um arquivo em disco com os rotulos originais, a correcao manual --
    que ja existe para as outras origens -- ficaria indisponivel justamente
    na origem com mais linhas nao reconhecidas.
    """
    destino = tmp_path / "weg.xlsx"
    dfs = importar_cvm(WEG, [2023, 2024], cache=DADOS, planilha=destino)

    # Qualquer linha publicada do balanco serve: o que se testa e o mecanismo de
    # remapear, e agora o candidato vem da arvore e nao da lista de sobras.
    alvo = _rotulo_publicado(dfs, "1.02.01")

    # O valor esperado vem da propria planilha, pela linha de mesmo rotulo.
    planilha = carregar_abas(destino)["Balanço"]
    coluna_2024 = list(planilha.iloc[0]).index(2024)
    esperado = float(planilha[planilha.iloc[:, 1] == alvo].iloc[0, coluna_2024])

    corrigido = aplicar_mapeamento_manual(dfs, destino, {alvo: "aplicacoes_financeiras"})

    assert corrigido.valor("aplicacoes_financeiras", 2024) == pytest.approx(esperado)
    assert "[manual]" in corrigido.mapeamento["aplicacoes_financeiras"]


def test_planilha_traz_os_valores_ja_em_reais(tmp_path):
    """A escala e resolvida antes de escrever: a planilha nao herda a pegadinha."""
    destino = tmp_path / "vivara.xlsx"
    importar_cvm(VIVARA, [2024], cache=DADOS, planilha=destino)

    dre = carregar_abas(destino)["DRE"]
    linha = dre[dre.iloc[:, 0] == "3.01"]
    assert float(linha.iloc[0, 2]) == pytest.approx(2_577_113_417.0)


def test_os_fixtures_continuam_sendo_recortes_e_nao_downloads():
    """Trava para o acidente que ja aconteceu nesta base.

    ``importar_cvm`` grava no diretorio de cache que recebe. Apontar o cache
    para ``tests/dados/cvm`` num script de exploracao faz o leitor baixar o zip
    anual inteiro -- 13 MB por ano -- exatamente ao lado dos recortes de 50 KB.
    Sem esta trava, isso entra num commit sem ninguem notar.
    """
    limite = 1_000_000
    pesados = [
        (arquivo.name, arquivo.stat().st_size)
        for arquivo in DADOS.iterdir()
        if arquivo.is_file() and arquivo.stat().st_size > limite
    ]
    assert not pesados, (
        f"arquivos acima de 1 MB em {DADOS}: {pesados}. Recorte antes de versionar "
        "-- os fixtures existentes tem ~50 KB."
    )


# ---------------------------------------------------------------------------
# Arrendamento reportado fora da subarvore de divida
# ---------------------------------------------------------------------------


def test_arrendamento_fora_da_divida_entra_na_divida(weg):
    """O defeito que esta funcao existe para corrigir, medido na WEG.

    O plano da CVM reserva 2.01.04.03 para arrendamento, dentro dos
    emprestimos. A WEG usa **outro** lugar, e la o passivo fica fora da divida:
    lido so pelo codigo fixo, o arrendamento aparecia como zero e a divida
    bruta saia menor do que e. Divida menor vira equity value maior, em
    silencio, porque a arvore publicada continua fechando.
    """
    arrendamento = weg.valor("arrendamento_curto_prazo", 2024) + weg.valor(
        "arrendamento_longo_prazo", 2024
    )
    assert arrendamento > 0, "o arrendamento da WEG voltou a sumir"
    divida = weg.divida_bruta()[2024]
    assert arrendamento < divida, "arrendamento nao pode exceder a divida de que faz parte"
    assert 0.10 < arrendamento / divida < 0.30


def test_o_leitor_avisa_quando_precisou_corrigir(weg):
    """Correcao silenciosa e pior que erro visivel: quem le confere na arvore."""
    assert any("fora da subárvore" in aviso for aviso in weg.avisos)


def test_linha_filha_nao_soma_duas_vezes():
    """Companhia que abre 'Arrendamentos' e, abaixo, 'Arrendamentos a pagar'."""
    from valuation.importacao.cvm import LinhaCVM, arrendamento_no_passivo

    def linha(codigo, descricao, valor):
        return LinhaCVM(
            codigo=codigo,
            descricao=descricao,
            valor=valor,
            ano=2024,
            demonstracao="bp",
            escala="MIL",
            escopo="con",
        )

    fora = arrendamento_no_passivo(
        [
            linha("2.02.02.02", "Arrendamentos", 100.0),
            linha("2.02.02.02.01", "Arrendamentos a pagar", 60.0),
            linha("2.02.02.02.02", "Arrendamento mercantil", 40.0),
        ]
    )
    assert fora["longo"][2024] == pytest.approx(100.0), "somou pai e filhas"
    assert fora["longo_fora"][2024] == pytest.approx(100.0)


def test_o_que_ja_esta_na_divida_nao_e_somado_de_novo():
    """2.01.04.03 ja entra por 2.01.04; soma-lo aqui contaria duas vezes."""
    from valuation.importacao.cvm import LinhaCVM, arrendamento_no_passivo

    fora = arrendamento_no_passivo(
        [
            LinhaCVM(
                codigo="2.01.04.03",
                descricao="Financiamento por Arrendamento",
                valor=500.0,
                ano=2024,
                demonstracao="bp",
                escala="MIL",
                escopo="con",
            )
        ]
    )
    # O arrendamento total conta a linha; a parcela "fora da divida", nao --
    # 2.01.04.03 ja entrou pela divida de curto prazo.
    assert fora["curto"][2024] == pytest.approx(500.0)
    assert not fora["curto_fora"] and not fora["longo_fora"]


# ---------------------------------------------------------------------------
# D&A: o erro que deixava 95% das companhias com EBITDA igual ao EBIT
# ---------------------------------------------------------------------------


def test_depreciacao_vem_da_dfc_quando_a_dre_nao_destaca(weg):
    """A WEG publica D&A so no ajuste da DFC, como quase toda companhia.

    Antes desta correcao ``depreciacao_dfc`` era **inalcancavel**: o rotulo
    "Depreciacao, Amortizacao e Exaustao" casava primeiro com a conta da DRE, e
    o filtro por demonstracao rejeitava sem tentar o proximo candidato. Medido
    em 150 companhias de 2024: **8 tinham D&A reconhecida, contra 116 depois**.
    Sem D&A, EBITDA = EBIT -- e a mediana da base tem D&A valendo 24% do EBIT.
    """
    assert weg.valor("depreciacao_amortizacao", 2024) == pytest.approx(812_485_000.0)
    assert weg.derivadas.get("depreciacao_amortizacao"), "veio da DRE, nao da DFC?"
    assert weg.mapeamento["depreciacao_dfc"].startswith("6.01.01.02")


def test_ebitda_da_weg_bate_com_o_publicado(weg):
    """EBITDA = EBIT + D&A, com os dois numeros da demonstracao real."""
    assert weg.ebitda()[2024] == pytest.approx(8_503_013_000.0)
    assert weg.ebitda()[2024] > weg.valor("ebit", 2024), "D&A sumiu de novo"
    margem = weg.ebitda()[2024] / weg.valor("receita_liquida", 2024)
    assert 0.20 < margem < 0.25


def test_plural_casa_com_o_sinonimo_singular():
    """A CVM escreve no plural e o vocabulario declarava no singular."""
    from valuation.importacao.esquema import reconhecer, singularizar

    assert singularizar("depreciacoes e amortizacoes") == "depreciacao e amortizacao"
    assert reconhecer("Depreciações e Amortizações", demonstracao="dre").chave == (
        "depreciacao_amortizacao"
    )
    # Palavra curta nao pode ser mutilada: "mais" nao vira "mal".
    assert singularizar("mais valia") == "mais valia"


def test_o_mesmo_rotulo_vai_para_a_conta_da_demonstracao_certa():
    """'Depreciacao e Amortizacao' e conta da DRE e ajuste da DFC."""
    from valuation.importacao.esquema import reconhecer

    assert reconhecer("Depreciação e Amortização", demonstracao="dre").chave == (
        "depreciacao_amortizacao"
    )
    assert reconhecer("Depreciação e Amortização", demonstracao="dfc").chave == (
        "depreciacao_dfc"
    )


def test_zero_na_dre_nao_bloqueia_a_derivacao():
    """Companhia que publica o campo de D&A como zero e o valor real na DFC."""
    import pandas as pd

    from valuation.importacao.esquema import DERIVACOES
    from valuation.importacao.importador import _derivar

    derivacao = next(d for d in DERIVACOES if d.chave == "depreciacao_amortizacao")
    assert derivacao.substitui_zero

    tabela = {
        "depreciacao_amortizacao": {2023: 0.0, 2024: 0.0},
        "depreciacao_dfc": {2023: 900.0, 2024: 1000.0},
    }
    _derivar(tabela, [2023, 2024])
    assert tabela["depreciacao_amortizacao"][2024] == pytest.approx(1000.0)


# ---------------------------------------------------------------------------
# Padronizacao do juro pago: FCO tem que significar a mesma coisa em todas
# ---------------------------------------------------------------------------


def _linha_dfc(codigo, descricao, valor, ano=2024):
    from valuation.importacao.cvm import LinhaCVM

    return LinhaCVM(
        codigo=codigo,
        descricao=descricao,
        valor=valor,
        ano=ano,
        demonstracao="dfc",
        escala="MIL",
        escopo="con",
    )


def test_juro_no_financiamento_e_identificado():
    from valuation.importacao.cvm import juros_pagos_no_financiamento

    total = juros_pagos_no_financiamento(
        [
            _linha_dfc("6.03.05", "Juros pagos", -500.0),
            _linha_dfc("6.03.06", "Amortização de empréstimos", -2000.0),
            _linha_dfc("6.01.02", "Juros pagos", -300.0),
        ]
    )
    # So o de 6.03 entra, e como magnitude.
    assert total == {2024: pytest.approx(500.0)}


def test_juro_recebido_e_jcp_nao_entram():
    """JCP e remuneracao ao acionista; juro recebido nao e custo da divida."""
    from valuation.importacao.cvm import juros_pagos_no_financiamento

    total = juros_pagos_no_financiamento(
        [
            _linha_dfc("6.03.05", "Juros recebidos", 200.0),
            _linha_dfc("6.03.06", "Juros sobre capital próprio pagos", -800.0),
            _linha_dfc("6.03.07", "Juros a pagar", -100.0),
        ]
    )
    assert total == {}


def test_juro_filho_nao_soma_duas_vezes():
    from valuation.importacao.cvm import juros_pagos_no_financiamento

    total = juros_pagos_no_financiamento(
        [
            _linha_dfc("6.03.05", "Juros pagos", -500.0),
            _linha_dfc("6.03.05.01", "Juros de empréstimos", -300.0),
            _linha_dfc("6.03.05.02", "Juros de debêntures", -200.0),
        ]
    )
    assert total == {2024: pytest.approx(500.0)}


def test_a_reclassificacao_preserva_a_identidade_da_dfc():
    """O que sai do financiamento entra no operacional: a soma nao muda.

    Medido na Petrobras de 2024, que classifica juro em financiamento: R$ 10,3
    bi saem do FCO. Sem a padronizacao, a conversao de caixa dela e a de uma
    companhia identica que classifica no operacional seriam numeros diferentes
    para a mesma economia -- e todo indicador que divide por FCO herdaria isso.
    """
    from valuation.importacao.cvm import _padronizar_juros_no_fco

    tabela = {
        "fluxo_operacional": {2024: 1000.0},
        "fluxo_financiamento": {2024: -600.0},
        "fluxo_investimento": {2024: -300.0},
    }
    avisos: list[str] = []
    soma_antes = sum(v[2024] for v in tabela.values())

    _padronizar_juros_no_fco(
        [_linha_dfc("6.03.05", "Juros pagos", -150.0)], tabela, {}, avisos
    )

    assert tabela["fluxo_operacional"][2024] == pytest.approx(850.0)
    assert tabela["fluxo_financiamento"][2024] == pytest.approx(-450.0)
    assert sum(v[2024] for v in tabela.values() if isinstance(v, dict) and 2024 in v) - tabela[
        "juros_pagos_no_financiamento"
    ][2024] == pytest.approx(soma_antes)
    assert avisos and "financiamento" in avisos[0]


def test_quem_ja_classifica_no_operacional_nao_e_tocado(weg):
    """A WEG ja poe o juro no operacional; mexer nela seria contar duas vezes."""
    import numpy as np

    movido = weg.valor("juros_pagos_no_financiamento", 2024)
    assert movido is None or not np.isfinite(movido)


def test_a_reclassificacao_aparece_como_conta_e_como_aviso():
    """Numero que muda sozinho, sem aparecer, e o pior tipo de correcao."""
    from valuation.importacao.cvm import _padronizar_juros_no_fco

    tabela = {
        "fluxo_operacional": {2024: 1000.0},
        "fluxo_financiamento": {2024: -600.0},
    }
    mapeamento: dict[str, str] = {}
    avisos: list[str] = []
    _padronizar_juros_no_fco(
        [_linha_dfc("6.03.05", "Juros pagos", -150.0)], tabela, mapeamento, avisos
    )

    assert tabela["juros_pagos_no_financiamento"][2024] == pytest.approx(150.0)
    assert "reclassificados" in mapeamento["fluxo_operacional"]
    assert "capital de giro" in avisos[0]


# ---------------------------------------------------------------------------
# Pagamento dentro do capital de giro, e outorga dentro da operacao
# ---------------------------------------------------------------------------


def test_pagamento_de_imposto_e_de_juro_sai_do_giro():
    """6.01.02 e "variacoes nos ativos e passivos" -- desembolso nao e variacao.

    Medido em 2024: 127 companhias lancam imposto de renda pago dentro do giro
    (R$ 44,7 bi) e 69 lancam juros pagos (R$ 23,4 bi). O FCO nao muda com isso,
    mas o investimento em giro que se le da DFC vira outra coisa -- e ele e
    premissa de projecao.
    """
    from valuation.importacao.cvm import pagamentos_dentro_do_giro

    total = pagamentos_dentro_do_giro(
        [
            _linha_dfc("6.01.02.03", "Imposto de renda e contribuição social pagos", -400.0),
            _linha_dfc("6.01.02.04", "Juros pagos", -150.0),
        ]
    )
    assert total == {2024: pytest.approx(550.0)}


def test_saldo_a_recuperar_e_a_recolher_continua_sendo_giro():
    """A separacao que importa e entre pagamento e movimento de saldo."""
    from valuation.importacao.cvm import pagamentos_dentro_do_giro

    total = pagamentos_dentro_do_giro(
        [
            _linha_dfc("6.01.02.05", "Impostos a recuperar", -120.0),
            _linha_dfc("6.01.02.06", "Tributos a recolher", 90.0),
            _linha_dfc("6.01.02.07", "Obrigações tributárias", 40.0),
        ]
    )
    assert total == {}


def test_tirar_pagamento_do_giro_nao_muda_o_fco():
    """Eles ja estavam dentro do FCO; so mudaram de linha."""
    from valuation.importacao.cvm import _reorganizar_o_fco

    tabela = {
        "variacao_capital_giro": {2024: -900.0},
        "fluxo_operacional": {2024: 1500.0},
        "fluxo_investimento": {2024: -400.0},
        "fluxo_financiamento": {2024: -200.0},
    }
    avisos: list[str] = []
    _reorganizar_o_fco(
        [_linha_dfc("6.01.02.03", "Imposto de renda e contribuição social pagos", -400.0)],
        tabela,
        {},
        avisos,
    )

    assert tabela["fluxo_operacional"][2024] == pytest.approx(1500.0), "o FCO mudou"
    assert tabela["variacao_capital_giro"][2024] == pytest.approx(-500.0)
    assert tabela["pagamentos_reclassificados_do_giro"][2024] == pytest.approx(400.0)
    assert avisos and "capital de giro" in avisos[0]


def test_outorga_sai_da_operacao_e_entra_no_investimento():
    """Comprar o direito de explorar e investimento, nao custo de operar."""
    from valuation.importacao.cvm import _reorganizar_o_fco

    tabela = {
        "fluxo_operacional": {2024: 1000.0},
        "fluxo_investimento": {2024: -500.0},
        "fluxo_financiamento": {2024: -300.0},
        "capex": {2024: 450.0},
    }
    soma_antes = 1000.0 - 500.0 - 300.0
    avisos: list[str] = []
    _reorganizar_o_fco(
        [_linha_dfc("6.01.03.04", "Pagamento de Obrigações com poder concedente", -145.0)],
        tabela,
        {},
        avisos,
    )

    assert tabela["fluxo_operacional"][2024] == pytest.approx(1145.0)
    assert tabela["fluxo_investimento"][2024] == pytest.approx(-645.0)
    # A identidade da DFC nao pode se mover.
    soma = (
        tabela["fluxo_operacional"][2024]
        + tabela["fluxo_investimento"][2024]
        + tabela["fluxo_financiamento"][2024]
    )
    assert soma == pytest.approx(soma_antes)
    # E a outorga vira capex, que e o que ela e.
    assert tabela["capex"][2024] == pytest.approx(595.0)
    assert tabela["outorga_paga"][2024] == pytest.approx(145.0)


def test_outorga_no_financiamento_tambem_vai_para_investimento():
    from valuation.importacao.cvm import _reorganizar_o_fco

    tabela = {
        "fluxo_operacional": {2024: 1000.0},
        "fluxo_investimento": {2024: -500.0},
        "fluxo_financiamento": {2024: -300.0},
    }
    _reorganizar_o_fco(
        [_linha_dfc("6.03.05", "Pagamento de Obrigações com poder concedente", -108.0)],
        tabela,
        {},
        [],
    )
    assert tabela["fluxo_financiamento"][2024] == pytest.approx(-192.0)
    assert tabela["fluxo_investimento"][2024] == pytest.approx(-608.0)


def test_opcoes_outorgadas_nao_sao_outorga_de_concessao():
    """O maior risco desta regra: "outorga" na DFC e quase sempre plano de opcoes.

    Medido: com padrao largo em "outorga" sao 38 companhias, e a maioria e
    remuneracao em acoes. Com o padrao estreito, 9 -- e todas sao concessao.
    """
    from valuation.importacao.cvm import outorgas_pagas

    fora = outorgas_pagas(
        [
            _linha_dfc("6.01.01.12", "Opções outorgadas reconhecidas", 30.0),
            _linha_dfc("6.01.01.13", "Despesa com outorga de opções", 12.0),
            _linha_dfc("6.01.01.14", "Instrumentos patrimoniais outorgados", 8.0),
            _linha_dfc("6.01.01.15", "Ações restritas outorgadas", 5.0),
        ]
    )
    assert fora == {"operacional": {}, "financiamento": {}}


def test_recebimento_do_poder_concedente_nao_e_pagamento_de_outorga():
    from valuation.importacao.cvm import outorgas_pagas

    fora = outorgas_pagas(
        [
            _linha_dfc("6.01.02.08", "Recebimento de Contas a Receber com o Poder Concedente", 448.0),
            _linha_dfc("6.01.01.10", "Obrigações e variação monetária com poder concedente", 275.0),
        ]
    )
    assert fora == {"operacional": {}, "financiamento": {}}


def test_a_weg_tem_pagamentos_dentro_do_giro(weg):
    """Caso real: o giro dela caiu de R$ 2.310 mi para R$ 774 mi consumidos."""
    movido = weg.valor("pagamentos_reclassificados_do_giro", 2024)
    assert movido == pytest.approx(1_535_663_000.0)
    assert any("capital de giro" in aviso for aviso in weg.avisos)


# ---------------------------------------------------------------------------
# O que parece juro pago e nao e
# ---------------------------------------------------------------------------


@pytest.mark.parametrize(
    "rotulo",
    [
        # Amortizacao de principal disfarcada: R$ 2,2 bi na Porto Seguro.
        "Pagamento de empréstimos e arrendamentos (exceto juros)",
        "Caixa líquido de custos financeiros, exceto juros",
        # Linha que mistura os dois: 18 linhas e R$ 21,5 bi na base.
        "Arrendamento - pagamentos de principal e juros",
        "Pagamento de juros e principal sobre passivo de arrendamento",
        # JCP e remuneracao ao acionista -- inclusive com a grafia errada que a
        # Dexxos publica.
        "Dividendos e Juros sobre capital prório pago a acionistas",
        "Juros sobre capital próprio pagos",
    ],
)
def test_o_que_nao_e_juro_pago_fica_de_fora(rotulo):
    from valuation.importacao.cvm import _NAO_E_JURO_PAGO

    assert _NAO_E_JURO_PAGO.search(rotulo), f"deveria excluir: {rotulo}"


@pytest.mark.parametrize(
    "rotulo",
    [
        "Juros pagos",
        "Juros pagos sobre empréstimos",
        "Juros sobre arrendamentos pagos",
        # AT1 de banco: e juro de verdade, e "capital principal" nao pode
        # confundi-lo com JCP.
        "Juros de instrumento elegível a capital principal pagos",
    ],
)
def test_juro_de_verdade_continua_contando(rotulo):
    from valuation.importacao.cvm import _MARCA_JUROS_PAGOS, _NAO_E_JURO_PAGO

    assert _MARCA_JUROS_PAGOS.search(rotulo)
    assert not _NAO_E_JURO_PAGO.search(rotulo), f"nao deveria excluir: {rotulo}"


def test_as_duas_leituras_do_juro_usam_o_mesmo_criterio():
    """A regra somada e a reclassificacao tem que concordar sobre o que e juro."""
    from valuation.importacao.cvm import _NAO_E_JURO_PAGO, REGRAS_SOMADAS

    regra = next(r for r in REGRAS_SOMADAS if r.chave == "juros_pagos")
    assert regra.exclui is _NAO_E_JURO_PAGO


def test_arrendamento_so_longo_nao_vaza_para_a_conta_de_curto():
    """A leitura do balanco manda nas duas contas, inclusive para apagar.

    Companhia que so publica arrendamento de longo prazo ficava com o
    reconhecimento por rotulo enchendo a conta de curto com o numero do longo --
    os rotulos sao identicos nos dois codigos. A auditoria achou dois casos em
    2023, e nenhuma identidade contabil os pegaria.
    """
    from valuation.importacao.cvm import LinhaCVM, _somar_arrendamento_fora_da_divida

    def linha(codigo, valor):
        return LinhaCVM(
            codigo=codigo,
            descricao="Financiamento por Arrendamento",
            valor=valor,
            ano=2024,
            demonstracao="bp",
            escala="MIL",
            escopo="con",
        )

    tabela = {
        # o reconhecimento por rotulo pos o numero do longo no curto
        "arrendamento_curto_prazo": {2024: 800.0},
        "divida_curto_prazo": {2024: 100.0},
        "divida_longo_prazo": {2024: 900.0},
    }
    mapeamento = {"arrendamento_curto_prazo": "2.02.01.03 - Financiamento por Arrendamento"}
    _somar_arrendamento_fora_da_divida([linha("2.02.01.03", 800.0)], tabela, mapeamento, [])

    assert "arrendamento_curto_prazo" not in tabela
    assert tabela["arrendamento_longo_prazo"][2024] == pytest.approx(800.0)


# ---------------------------------------------------------------------------
# O app abre todas as demonstracoes, e nao so as do modelo
# ---------------------------------------------------------------------------


def test_a_arvore_traz_as_seis_demonstracoes(weg):
    """Nada do que a companhia publica fica de fora da arvore.

    O vocabulario canonico nomeia umas dezenas de contas porque e delas que o
    motor precisa. O resto **nao e descartado**: fica na arvore publicada, e
    quem usa escolhe o nivel de abertura. Antes de abrir DVA, DMPL e DRA, mais
    da metade do que a WEG publica ficava fora -- 298 de 574 linhas.
    """
    presentes = set(weg.detalhe["demonstracao"])
    assert {"dre", "bp", "dfc", "dva", "dra", "dmpl"} <= presentes


def test_a_dva_responde_o_que_a_dre_nao_abre(weg):
    """Receita bruta, folha e impostos totais nao existem na DRE padronizada."""
    bruta = weg.valor("receita_bruta", 2024)
    liquida = weg.valor("receita_liquida", 2024)

    assert bruta > liquida, "a receita bruta tem que superar a liquida"
    # A diferenca sao impostos sobre vendas e devolucoes.
    assert 0.03 < (bruta - liquida) / bruta < 0.35
    assert weg.valor("pessoal", 2024) > 0
    assert weg.valor("impostos_taxas_contribuicoes", 2024) > 0


def test_a_dmpl_entra_somada_pelas_colunas_do_patrimonio():
    """A mesma conta aparece uma vez por componente do PL; a arvore soma.

    Sem somar, a arvore teria o mesmo codigo repetido cinco ou seis vezes e
    ninguem conseguiria ler. Quem precisa da abertura por componente vai ao
    arquivo -- e o vocabulario nao promete tê-la.
    """
    from valuation.importacao.cvm import _DEMONSTRACOES_COM_COLUNA

    assert "dmpl" in _DEMONSTRACOES_COM_COLUNA


def test_nenhuma_demonstracao_nova_atrapalha_o_reconhecimento(weg):
    """Codigos da DVA (7.x) e da DRA (4.x) nao podem invadir contas da DRE."""
    for chave in ("receita_liquida", "ebit", "lucro_liquido"):
        origem = weg.mapeamento.get(chave, "")
        assert origem.startswith("3."), f"{chave} veio de {origem}"


# ---------------------------------------------------------------------------
# O sinal do imposto vem da identidade, e nao de convencao de fonte
# ---------------------------------------------------------------------------


def test_credito_de_imposto_nao_vira_despesa(catalogo):
    """118 das 467 companhias de 2024 publicam ``3.08`` positivo -- R$ 71 bi.

    Guardado como magnitude, o credito virava despesa. O estrago nao aparecia na
    DRE (que le ``3.08.01``/``3.08.02``) e sim na **aliquota efetiva**: ela e
    ``impostos / LAIR`` clipada em [0, 1], entao credito lido como despesa sobe
    a aliquota em vez de zera-la, e com credito grande sobre LAIR pequeno ela
    bate 100% e **zera o NOPAT**.
    """
    dfs = importar_cvm(24805, [2024], cache=DADOS, catalogo=catalogo)
    publicado = _linhas_publicadas(24805, "con", 2024)["3.08"]
    assert publicado > 0, "o fixture deixou de ser um caso de credito"

    assert dfs.valor("impostos", 2024) == pytest.approx(-publicado)
    assert "impostos" in dfs.derivadas
    assert "credito" in dfs.derivadas["impostos"]


def test_despesa_de_imposto_continua_positiva(catalogo):
    """A convencao nao mudou: despesa positiva. So o credito ganhou sinal."""
    weg = importar_cvm(5410, [2024], cache=DADOS, catalogo=catalogo)
    assert weg.valor("impostos", 2024) > 0
    assert "impostos" not in weg.derivadas


def test_a_correcao_de_sinal_nao_plugga_diferenca_de_valor():
    """Corrige o sinal, nunca o valor -- plug esconderia erro de leitura."""
    from valuation.importacao.importador import _corrigir_sinal_dos_impostos

    # Identidade diz -100 (credito), mas a conta lida tem outra magnitude.
    tabela = {
        "impostos": {2024: 250.0},
        "lucro_antes_impostos": {2024: 400.0},
        "lucro_liquido": {2024: 500.0},
    }
    assert _corrigir_sinal_dos_impostos(tabela, [2024]) == ""
    assert tabela["impostos"][2024] == 250.0

    # Mesma magnitude, sinal trocado: aí sim corrige.
    tabela["impostos"] = {2024: 100.0}
    assert _corrigir_sinal_dos_impostos(tabela, [2024]) == "2024"
    assert tabela["impostos"][2024] == -100.0


def test_a_operacao_descontinuada_sai_antes_da_identidade():
    """O imposto so alcanca a operacao continuada; 3.11 traz as duas."""
    from valuation.importacao.importador import _corrigir_sinal_dos_impostos

    tabela = {
        "impostos": {2024: 40.0},          # magnitude de um credito
        "lucro_antes_impostos": {2024: 100.0},
        "lucro_liquido": {2024: 170.0},    # 140 continuadas + 30 descontinuadas
        "operacoes_descontinuadas": {2024: 30.0},
    }
    assert _corrigir_sinal_dos_impostos(tabela, [2024]) == "2024"
    assert tabela["impostos"][2024] == -40.0


def test_corrente_e_diferido_guardam_o_sinal_publicado(catalogo):
    """Cada um pode ser credito por conta propria, e em geral discordam.

    Medido no DFP consolidado de 2024: 221 das 467 companhias tiveram credito no
    **diferido**, 16 no corrente, 8 nas duas, e em **204 os dois tem sinais
    opostos** -- o caso mais comum da base. Conferido conta a conta contra o
    arquivo em 449 das 467; as 18 restantes sao banco ou seguradora, que publicam
    em outro plano de contas e ja sao detectadas.
    """
    for codigo in (5410, 24805):
        dfs = importar_cvm(codigo, [2024], cache=DADOS, catalogo=catalogo)
        publicadas = _linhas_publicadas(codigo, "con", 2024)
        for chave, cod in (
            ("imposto_corrente", "3.08.01"),
            ("imposto_diferido", "3.08.02"),
        ):
            if cod not in publicadas:
                continue
            assert dfs.valor(chave, 2024) == pytest.approx(publicadas[cod]), (
                f"{chave} de {codigo} perdeu o sinal publicado"
            )


def test_a_soma_das_duas_reconstroi_o_total(catalogo):
    """``3.08.01 + 3.08.02 = 3.08`` fecha com o sinal publicado, nunca com magnitude.

    Sao 440 de 440 companhias que abrem as duas contas. Como ``impostos`` guarda
    despesa positiva e as filhas o sinal publicado, a soma delas e ``-impostos``.
    """
    for codigo in (5410, 24805):
        dfs = importar_cvm(codigo, [2024], cache=DADOS, catalogo=catalogo)
        soma = dfs.valor("imposto_corrente", 2024) + dfs.valor("imposto_diferido", 2024)
        assert soma == pytest.approx(-dfs.valor("impostos", 2024), rel=1e-9)


def test_planilha_com_as_duas_em_magnitude_tem_o_sinal_recuperado():
    """O template mandava digitar magnitude; quem seguiu tem conserto.

    Quando a soma tem a magnitude do total e o sinal oposto, foi magnitude, e a
    identidade devolve os dois ao lado certo.
    """
    from valuation.importacao.importador import _corrigir_sinal_do_ir_aberto

    tabela = {
        "imposto_corrente": {2024: 60.0},   # digitados como despesa positiva
        "imposto_diferido": {2024: 40.0},
        "impostos": {2024: 100.0},          # despesa positiva, convencao do app
    }
    assert _corrigir_sinal_do_ir_aberto(tabela, [2024]) == "2024"
    assert tabela["imposto_corrente"][2024] == -60.0
    assert tabela["imposto_diferido"][2024] == -40.0


def test_sinais_opostos_digitados_como_magnitude_nao_sao_inventados():
    """A informacao se perdeu na origem, e nenhuma identidade a recupera.

    Corrente de -60 com diferido de +20 da total de -40. Digitados como 60 e 20,
    a soma e 80: magnitude nenhuma bate com 40, e o certo e **nao mexer**.
    """
    from valuation.importacao.importador import _corrigir_sinal_do_ir_aberto

    tabela = {
        "imposto_corrente": {2024: 60.0},
        "imposto_diferido": {2024: 20.0},
        "impostos": {2024: 40.0},
    }
    assert _corrigir_sinal_do_ir_aberto(tabela, [2024]) == ""
    assert tabela["imposto_corrente"][2024] == 60.0
    assert tabela["imposto_diferido"][2024] == 20.0


def test_quando_nao_da_para_recuperar_o_app_avisa():
    """Numero que nao fecha e nao aparece e o pior tipo de erro."""
    import pandas as pd

    from valuation.importacao.importador import _conferir

    valores = pd.DataFrame(
        {2024: {"imposto_corrente": 60.0, "imposto_diferido": 20.0, "impostos": 40.0}}
    )
    avisos: list[str] = []
    _conferir(valores, avisos)
    assert any("nao reconstroem o IR total" in a for a in avisos), avisos


def test_o_template_nao_promete_magnitude_para_o_ir(tmp_path):
    """A instrucao antiga levava o usuario direto ao erro que acabamos de corrigir."""
    from openpyxl import load_workbook

    from valuation.importacao.template import gerar_template

    caminho = gerar_template(tmp_path / "modelo.xlsx")
    aba = load_workbook(caminho)["Instrucoes"]
    texto = " ".join(
        str(c.value) for (c,) in aba.iter_rows(min_col=1, max_col=1) if c.value
    )
    assert "IR e CSLL correntes" in texto and "IR e CSLL diferidos" in texto
    assert "despesa negativa, credito positivo" in texto
    # E a regra da magnitude nao pode mais citar impostos.
    linha_magnitude = next(
        linha
        for linha in texto.split(".")
        if "magnitude e padroniza o sinal" in linha
    )
    assert "imposto" not in linha_magnitude.lower()


def test_companhia_sem_minoritario_zera_as_filhas_de_3_11(catalogo):
    """102 das 467 publicam ``3.11.01 = 0`` e ``3.11.02 = 0`` com ``3.11`` != 0.

    Nao e que os controladores nao tenham ganhado nada -- e que a companhia nao
    tem minoritario e nao se deu ao trabalho de repetir o total na filha. Lido ao
    pe da letra, o lucro dos controladores da CESP seria zero em vez dos
    R$ 1.077,9 mi que ela ganhou.

    Mesmo caso da D&A: zero publicado que quer dizer "nao abri", nao "nao tem".
    """
    from valuation.importacao.esquema import POR_CHAVE  # noqa: F401
    from valuation.importacao.importador import _derivar

    tabela = {
        "lucro_liquido": {2024: 1_077_900_000.0},
        "lucro_controladores": {2024: 0.0},
        "lucro_nao_controladores": {2024: 0.0},
    }
    derivadas = _derivar(tabela, [2024])
    assert tabela["lucro_controladores"][2024] == 1_077_900_000.0
    assert "lucro_controladores" in derivadas


def test_o_zero_nao_atropela_quem_publica_a_abertura():
    """Quem abre 3.11.01 de verdade nao pode ter o valor substituido."""
    from valuation.importacao.importador import _derivar

    tabela = {
        "lucro_liquido": {2024: 82_440_000.0},
        "lucro_controladores": {2024: 79_514_000.0},
        "lucro_nao_controladores": {2024: 0.0},
    }
    _derivar(tabela, [2024])
    assert tabela["lucro_controladores"][2024] == 79_514_000.0


def test_companhia_com_tudo_zerado_continua_zerada():
    """Rio Paranapanema e TIM S.A. publicam 3.11 zero. Zero ali e zero mesmo."""
    from valuation.importacao.importador import _derivar

    tabela = {
        "lucro_liquido": {2024: 0.0},
        "lucro_controladores": {2024: 0.0},
        "lucro_nao_controladores": {2024: 0.0},
    }
    _derivar(tabela, [2024])
    assert tabela["lucro_controladores"][2024] == 0.0


# ---------------------------------------------------------------------------
# A D&A do EBITDA vem da DFC, nao da DRE
# ---------------------------------------------------------------------------


def test_a_da_da_dre_e_so_o_pedaco_do_sga():
    """``3.04.02.x`` mora dentro de "Despesas Gerais e Administrativas".

    Ela captura a depreciacao que correu pelo SG&A e nao a que correu pelo CPV --
    que numa industria ou concessionaria e a maior parte. O ajuste da DFC devolve
    ao lucro **toda** a D&A que o reduziu, que e o que ``EBITDA = EBIT + D&A``
    pede. Medido nas 467 companhias de 2024: entre as 56 que publicam as duas, a
    da DFC **nunca e menor** -- 34 coincidem e em 22 a da DFC e maior, ate 310x.
    """
    from valuation.importacao.importador import _preferir_a_da_do_fluxo_de_caixa

    # CPFL Energia de 2024, em reais.
    tabela = {
        "depreciacao_amortizacao": {2024: 142_031_000.0},   # 3.04.02.01
        "depreciacao_dfc": {2024: 2_303_124_000.0},         # 6.01.01.02
    }
    assert _preferir_a_da_do_fluxo_de_caixa(tabela, [2024]) == "2024"
    assert tabela["depreciacao_amortizacao"][2024] == 2_303_124_000.0


def test_quando_as_duas_coincidem_nada_e_registrado():
    """34 das 56 ja coincidem; registrar troca ali so faria ruido na tela."""
    from valuation.importacao.importador import _preferir_a_da_do_fluxo_de_caixa

    tabela = {
        "depreciacao_amortizacao": {2024: 812_485_000.0},
        "depreciacao_dfc": {2024: 812_485_000.0},
    }
    assert _preferir_a_da_do_fluxo_de_caixa(tabela, [2024]) == ""


def test_sem_da_na_dfc_a_da_dre_fica():
    """5 companhias so tem a linha da DRE. Trocar por nada seria perder o numero."""
    from valuation.importacao.importador import _preferir_a_da_do_fluxo_de_caixa

    tabela = {
        "depreciacao_amortizacao": {2024: 500.0},
        "depreciacao_dfc": {2024: 0.0},
    }
    assert _preferir_a_da_do_fluxo_de_caixa(tabela, [2024]) == ""
    assert tabela["depreciacao_amortizacao"][2024] == 500.0


def test_a_troca_aparece_como_conta_derivada(catalogo):
    """Numero que muda sozinho sem aparecer e o pior tipo de correcao."""
    dfs = importar_cvm(5410, [2024], cache=DADOS, catalogo=catalogo)
    assert "depreciacao_amortizacao" in dfs.derivadas
    assert "DFC" in dfs.derivadas["depreciacao_amortizacao"]


def test_o_ebitda_da_weg_nao_muda_com_a_troca(catalogo):
    """A WEG ja vinha da DFC; a mudanca de prioridade nao pode mexer nela."""
    dfs = importar_cvm(5410, [2024], cache=DADOS, catalogo=catalogo)
    assert dfs.ebitda()[2024] == pytest.approx(8_503_013_000.0)


def test_o_de_para_aponta_para_a_linha_que_virou_o_numero(catalogo):
    """Origem registrada errada e erro que soma nenhuma denuncia.

    Trocar o valor e deixar o mapeamento apontando para a linha da DRE faria a
    auditoria de origem reportar ``3.04.02.x`` para um numero que veio de
    ``6.01.01.x``. E exatamente o tipo de coisa que ela existe para pegar.
    """
    dfs = importar_cvm(5410, [2024], cache=DADOS, catalogo=catalogo)
    origem = dfs.mapeamento.get("depreciacao_amortizacao", "")
    if "depreciacao_amortizacao" in dfs.derivadas and origem:
        assert origem.startswith("6."), origem


# ---------------------------------------------------------------------------
# A DFC pelo metodo direto usa os mesmos codigos para outras contas
# ---------------------------------------------------------------------------

DIRETA = 3328  # publica a DFC pelo metodo direto


def test_o_metodo_direto_e_detectado_pelo_arquivo():
    """A CVM publica os dois metodos em arquivos separados, e isso e a verdade.

    Pelo rotulo nao daria: das 16 companhias de 2024 no metodo direto, **so 9
    abrem com "Recebimento de Consumidores"**. O arquivo (``DFC_MD`` contra
    ``DFC_MI``) e a declaracao da propria companhia.
    """
    dfs = importar_cvm(DIRETA, [2024], cache=DADOS)
    assert any("metodo direto" in a for a in dfs.avisos), dfs.avisos


def test_as_contas_que_so_existem_no_indireto_ficam_em_branco():
    """No direto, 6.01.01 e "Recebimento de Consumidores".

    Ler pelo codigo poria recebimento de clientes em ``caixa_das_operacoes`` e
    fornecedores em ``variacao_capital_giro``. Numero errado na conta certa,
    calado -- e era a causa das 5 unicas companhias em que a decomposicao do FCO
    nao fechava.
    """
    dfs = importar_cvm(DIRETA, [2024], cache=DADOS)
    for chave in ("caixa_das_operacoes", "variacao_capital_giro", "depreciacao_dfc"):
        if chave in dfs.valores.index:
            assert not np.isfinite(dfs.valor(chave, 2024)), (
                f"{chave} nao existe na DFC pelo metodo direto"
            )


def test_o_que_e_igual_nos_dois_metodos_continua_lido():
    """So 6.01.xx muda de significado. Os totais de secao sao os mesmos."""
    dfs = importar_cvm(DIRETA, [2024], cache=DADOS)
    assert np.isfinite(dfs.valor("fluxo_operacional", 2024))
    assert np.isfinite(dfs.valor("fluxo_investimento", 2024))
    assert np.isfinite(dfs.valor("capex", 2024))


def test_o_metodo_indireto_nao_foi_afetado(catalogo):
    """A WEG e as demais do fixture continuam lendo tudo."""
    weg = importar_cvm(5410, [2024], cache=DADOS, catalogo=catalogo)
    assert not any("metodo direto" in a for a in weg.avisos)
    assert np.isfinite(weg.valor("caixa_das_operacoes", 2024))
    assert np.isfinite(weg.valor("variacao_capital_giro", 2024))
    assert np.isfinite(weg.valor("depreciacao_amortizacao", 2024))


def test_a_regra_somada_da_da_nao_varre_o_direto():
    """``6.01.01.`` no indireto e a secao de ajustes ao lucro; no direto, nao.

    A regra que soma D&A varre aquele prefixo sem exigir verbo, porque ali linha
    que fala de depreciacao **e** depreciacao. No metodo direto a mesma varredura
    pegaria linha de recebimento com "amortizacao" no nome.
    """
    dfs = importar_cvm(DIRETA, [2024], cache=DADOS)
    assert "depreciacao_dfc" not in dfs.mapeamento


# ---------------------------------------------------------------------------
# O plano financeiro tambem nao e um so
# ---------------------------------------------------------------------------


def test_o_codigo_do_plano_financeiro_so_vale_com_o_aval_do_rotulo():
    """``2.07`` é o patrimônio líquido em 10 companhias e outra coisa em 7.

    As 7 usam o layout com passivos abertos por critério de mensuração IFRS 9 —
    Itaú, BTG, Pine — onde ``2.07`` é "Passivos sobre Ativos Não Correntes a
    Venda" e o patrimônio está em ``2.08``. Confiar só no código punha **zero**
    no patrimônio líquido do maior banco do país, e nenhuma identidade
    denunciava, porque zero é um número tão válido quanto qualquer outro.
    """
    from valuation.importacao.cvm import (
        CODIGOS_PLANO_FINANCEIRO,
        LinhaCVM,
        PLANO_FINANCEIRO,
        _reconhecer_na_demonstracao,
    )

    assert CODIGOS_PLANO_FINANCEIRO["2.07"] == "patrimonio_liquido"

    def linha(codigo: str, descricao: str) -> LinhaCVM:
        return LinhaCVM(
            codigo=codigo,
            descricao=descricao,
            valor=1.0,
            ano=2024,
            demonstracao="bp",
            escala="MIL",
            escopo="con",
        )

    # Layout em que o código está certo: o rótulo confirma.
    concorda = _reconhecer_na_demonstracao(
        linha("2.07", "Patrimônio Líquido Consolidado"), PLANO_FINANCEIRO
    )
    assert concorda.chave == "patrimonio_liquido"

    # Layout em que não está: o rótulo não avaliza, e a linha fica sem conta.
    discorda = _reconhecer_na_demonstracao(
        linha("2.07", "Passivos sobre Ativos Não Correntes a Venda e Descontinuados"),
        PLANO_FINANCEIRO,
    )
    assert discorda.chave != "patrimonio_liquido"

    # E o 2.08 daquele layout é alcançado pelo rótulo.
    outro = _reconhecer_na_demonstracao(
        linha("2.08", "Patrimônio Líquido Consolidado"), PLANO_FINANCEIRO
    )
    assert outro.chave == "patrimonio_liquido"


def test_o_vocabulario_alcanca_os_rotulos_de_banco():
    """Sem eles o aval do rótulo derrubaria a cobertura em vez de corrigi-la.

    Medido nas 20 companhias do plano financeiro de 2024: 17 escrevem
    "Receitas de/da Intermediação Financeira" onde a indústria escreve "Receita
    de Venda de Bens e/ou Serviços", e 10 fecham a DRE com "Lucro ou Prejuízo
    Líquido Consolidado do Período".
    """
    from valuation.importacao.esquema import reconhecer

    esperado = {
        ("Receitas de Intermediação Financeira", "dre"): "receita_liquida",
        ("Receitas da Intermediação Financeira", "dre"): "receita_liquida",
        ("Despesas de Intermediação Financeira", "dre"): "custo_produtos_vendidos",
        ("Resultado Bruto de Intermediação Financeira", "dre"): "lucro_bruto",
        ("Lucro ou Prejuízo Líquido Consolidado do Período", "dre"): "lucro_liquido",
    }
    for (rotulo, demonstracao), chave in esperado.items():
        assert reconhecer(rotulo, None, demonstracao).chave == chave, rotulo


def test_linha_publicada_duas_vezes_nao_conta_duas_vezes():
    """A CVM publica linhas **idênticas em todos os campos** para 2 companhias.

    Não é versão do documento nem período diferente: é a mesma linha duas vezes,
    byte a byte. Medido no DFP consolidado de 2024, são o Grupo Salta (662
    linhas) e a CPX Distribuidora (626), espalhadas por todas as demonstrações.

    O estrago é seletivo, e por isso passou despercebido: conta reconhecida por
    código único não muda — a segunda leitura sobrescreve a primeira com o mesmo
    número —, mas **regra somada conta as duas**. Nenhuma identidade denuncia,
    porque as seções da DFC dobram todas na mesma proporção.
    """
    import pandas as pd

    from valuation.importacao.cvm import _sem_linhas_repetidas

    linha = {"CD_CONTA": "6.01.03.02", "VL_CONTA": "-207306.0", "ORDEM_EXERC": "ÚLTIMO"}
    repetido = pd.DataFrame([linha, linha, {**linha, "CD_CONTA": "6.01.03.01"}])

    limpo = _sem_linhas_repetidas(repetido, "dfc")
    assert len(limpo) == 2

    # A DMPL fica de fora: nela a mesma conta aparece uma vez por componente do
    # patrimônio, e é a soma por COLUNA_DF que resolve.
    assert len(_sem_linhas_repetidas(repetido, "dmpl")) == 3


def test_o_arquivo_decide_o_metodo_da_dfc_e_o_rotulo_nao_o_sobrepoe():
    """Uma linha de giro não transforma uma DFC indireta em direta.

    `detectar_metodo_da_dfc` tinha uma rede por rótulo que a própria docstring
    dizia existir "só para origens que não são o zip da CVM" — mas ela rodava
    também na CVM, **por cima** da declaração da companhia. Medido em 2024, três
    companhias publicam só no `DFC_MI` e tinham uma linha de capital de giro que
    o padrão lia como recebimento de clientes:

        Americanas       "Adiantamentos recebidos de clientes"
        Vamos Locação    "Juros recebidos de clientes"
        Bioma Educação   "Perdas nos recebimentos de clientes"

    Nas três o app anunciava método direto — falso — e descartava as quatro
    contas que só existem no indireto, entre elas a D&A: R$ 1.010,0 mi na
    Americanas, que estava no arquivo e não era lida. Sem D&A o EBITDA era o
    próprio EBIT.
    """
    from valuation.importacao.cvm import LinhaCVM, detectar_metodo_da_dfc

    def linha(codigo, descricao, grupo):
        return LinhaCVM(
            codigo=codigo,
            descricao=descricao,
            valor=1.0,
            ano=2024,
            demonstracao="dfc",
            escala="MIL",
            escopo="con",
            grupo=grupo,
        )

    # So no arquivo do indireto, com o rotulo tentador: e indireto.
    do_indireto = [
        linha("6.01.01.03", "Depreciação e Amortização", "DFC_MI"),
        linha("6.01.02.04", "Adiantamentos recebidos de clientes", "DFC_MI"),
    ]
    assert detectar_metodo_da_dfc(do_indireto) == "indireto"

    # No arquivo do direto: e direto, mesmo sem rotulo caracteristico -- das 16
    # companhias de 2024, so 9 abrem com "Recebimento de Consumidores".
    do_direto = [linha("6.01.01", "Caixa recebido de terceiros", "DFC_MD")]
    assert detectar_metodo_da_dfc(do_direto) == "direto"

    # Sem arquivo (planilha importada), a rede por rotulo continua valendo: e a
    # unica pista que existe ali.
    sem_arquivo = [
        LinhaCVM(
            codigo="6.01.01",
            descricao="Recebimento de Consumidores",
            valor=1.0,
            ano=2024,
            demonstracao="dfc",
            escala="MIL",
            escopo="con",
        )
    ]
    assert detectar_metodo_da_dfc(sem_arquivo) == "direto"


def test_consolidado_zerado_nao_e_escopo_valido(tmp_path):
    """Linha existir não é o mesmo que ter dado.

    `escopo_da_companhia` decidia por "existe alguma linha da companhia neste
    escopo", e há companhia que entrega o consolidado com o plano de contas
    inteiro **zerado**. O app lia zeros e montava uma companhia vazia, sem cair
    no individual — a queda só acontecia quando o consolidado faltava.

    A TIM S.A. é o caso: a DFP consolidada dela é zero desde o exercício de 2024
    (era R$ 23.833,9 mi em 2023) e a individual traz R$ 25.447,9 mi. A
    demonstração publicada pela companhia **tem** o consolidado; o que vem vazio
    é o extrato estruturado desse escopo, e o ITR consolidado segue cheio.
    """
    import zipfile

    from valuation.importacao.cvm import escopo_da_companhia

    # O layout e o do arquivo real: o recorte em bytes acha CD_CVM pela
    # **posicao** da coluna, entao um CSV inventado com outra ordem nao e lido.
    cabecalho = (
        "CNPJ_CIA;DT_REFER;VERSAO;DENOM_CIA;CD_CVM;GRUPO_DFP;MOEDA;ESCALA_MOEDA;"
        "ORDEM_EXERC;DT_INI_EXERC;DT_FIM_EXERC;CD_CONTA;DS_CONTA;VL_CONTA;ST_CONTA_FIXA"
    )

    def csv(valor):
        linha = (
            "02.421.421/0001-11;2024-12-31;1;TIM S.A.;024929;DF Consolidado;REAL;"
            "MIL;ÚLTIMO;2024-01-01;2024-12-31;3.01;Receita;" + valor + ";S"
        )
        return (cabecalho + "\n" + linha + "\n").encode("latin-1")

    caminho = tmp_path / "dfp.zip"

    def montar(con, ind):
        with zipfile.ZipFile(caminho, "w") as z:
            z.writestr("dfp_cia_aberta_DRE_con_2024.csv", csv(con))
            z.writestr("dfp_cia_aberta_DRE_ind_2024.csv", csv(ind))

    # O consolidado existe, mas nao tem numero: vale o individual.
    montar(con="0", ind="25447900")
    assert escopo_da_companhia(caminho, 2024, 24929) == "ind"

    # Com numero nos dois, o consolidado continua ganhando.
    montar(con="23833900", ind="23843000")
    assert escopo_da_companhia(caminho, 2024, 24929) == "con"

    # Zerado nos dois: continua devolvendo um escopo, e nao `None` -- "sem DFP"
    # seria outra afirmacao, e falsa.
    montar(con="0", ind="0")
    assert escopo_da_companhia(caminho, 2024, 24929) == "con"


def test_da_dentro_do_nao_recorrente_vira_aviso():
    """Amortização em "outras despesas" infla a margem recorrente, e ela é a base.

    `itens_nao_recorrentes` é `3.04.03 + 3.04.04 + 3.04.05` com o sinal
    publicado, e a **margem EBITDA recorrente é o que `sugerir_premissas`
    projeta**. Amortização de intangível é a coisa mais recorrente que existe:
    onde a companhia a lança nesse bloco, a margem recorrente sai inflada e a
    projeção parte dela.

    Medido no DFP consolidado de 2024: 43 linhas em 41 companhias, e não são
    pequenas — CBD com R$ 1.045,0 mi (239,7% do EBIT), Casas Bahia com R$ 864,0
    mi, Marisa com R$ 166,4 mi (382,2%).

    O app **avisa e não corrige**: excluir a linha mudaria a base de projeção de
    41 companhias em silêncio, e amortização de mais-valia é caso legitimamente
    discutível.
    """
    from valuation.importacao.cvm import (
        LinhaCVM,
        _avisar_da_dentro_do_nao_recorrente,
    )

    def linha(codigo, descricao):
        return LinhaCVM(
            codigo=codigo,
            descricao=descricao,
            valor=-164_300_000.0,
            ano=2024,
            demonstracao="dre",
            escala="MIL",
            escopo="con",
        )

    avisos: list[str] = []
    _avisar_da_dentro_do_nao_recorrente(
        [linha("3.04.05.01", "Amortização de intangíveis")], avisos
    )
    assert len(avisos) == 1
    assert "164,3 milhões" in avisos[0], avisos[0]
    assert "Amortização de intangíveis" in avisos[0]

    # Fora do bloco nao recorrente, nao ha o que avisar: ali a D&A esta no lugar.
    avisos = []
    _avisar_da_dentro_do_nao_recorrente(
        [linha("3.04.02.06", "Despesa de Depreciação")], avisos
    )
    assert avisos == []

    # E uma linha do bloco que nao fala de D&A tambem nao acusa.
    avisos = []
    _avisar_da_dentro_do_nao_recorrente(
        [linha("3.04.05.02", "Perdas com processos judiciais")], avisos
    )
    assert avisos == []


def test_a_da_da_dva_e_ultimo_recurso_e_nao_substitui_a_da_dfc():
    """`7.04.01` preenche o buraco, e nunca troca o número de quem já tem.

    A DVA declara "Depreciação, Amortização e Exaustão" num código padronizado, e
    há companhia que só a publica ali — a Farmácia e Drogaria Nissei tem R$ 104,6
    mi só na DVA, e a Axia Energia Nordeste R$ 633,6 mi sobre R$ 8,0 bi de
    receita (uma transmissora sem depreciação, que não existe).

    **Mas ela não é equivalente à da DFC**, e é a medição que impõe isso: nas 422
    companhias de 2024 que publicam as duas, concordam em 328 e **discordam em
    94** — Rumo com 2.303,4 contra 5.452,6, São Martinho com 1.150,0 contra
    2.348,4. A DVA carrega exaustão e mede em base própria.
    """
    from valuation.importacao.importador import _da_da_dva_como_ultimo_recurso

    # Ja tem D&A: a DVA nao encosta, mesmo discordando.
    tabela = {
        "depreciacao_amortizacao": {2024: 2_303_400_000.0},
        "depreciacao_dva": {2024: -5_452_600_000.0},
    }
    assert _da_da_dva_como_ultimo_recurso(tabela, [2024]) == ""
    assert tabela["depreciacao_amortizacao"][2024] == 2_303_400_000.0

    # Nao tem: a DVA entra, em magnitude (a DVA publica a retencao negativa).
    tabela = {"depreciacao_dva": {2024: -104_600_000.0}}
    assert _da_da_dva_como_ultimo_recurso(tabela, [2024]) == "2024"
    assert tabela["depreciacao_amortizacao"][2024] == 104_600_000.0

    # Zero publicado tambem e buraco: zero nao e D&A de uma operacao.
    tabela = {
        "depreciacao_amortizacao": {2024: 0.0},
        "depreciacao_dva": {2024: -633_600_000.0},
    }
    assert _da_da_dva_como_ultimo_recurso(tabela, [2024]) == "2024"
    assert tabela["depreciacao_amortizacao"][2024] == 633_600_000.0

    # Sem DVA nenhuma, nada acontece.
    tabela = {"depreciacao_amortizacao": {2024: 10.0}}
    assert _da_da_dva_como_ultimo_recurso(tabela, [2024]) == ""

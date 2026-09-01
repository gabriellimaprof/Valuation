"""Testes da importacao de demonstracoes financeiras.

Cada teste monta uma planilha no formato de uma origem real (template proprio,
export CVM/B3, export de terminal) e confere que as tres chegam ao mesmo
vocabulario canonico.
"""

from __future__ import annotations

import numpy as np
import pandas as pd
import pytest

from valuation.importacao import (
    aplicar_mapeamento_manual,
    gerar_template,
    importar,
    normalizar,
    para_numero,
    reconhecer,
)
from valuation.importacao.leitura import extrair_ano, localizar_grade

# ---------------------------------------------------------------------------
# Leitura de numeros e cabecalhos
# ---------------------------------------------------------------------------


@pytest.mark.parametrize(
    "entrada,esperado",
    [
        (1234.5, 1234.5),
        ("1234.5", 1234.5),
        ("1.234.567,89", 1234567.89),      # formato brasileiro
        ("1,234,567.89", 1234567.89),      # formato americano
        ("1.234", 1234.0),                 # milhar brasileiro sem decimal
        ("1,5", 1.5),                      # decimal brasileiro
        ("(123)", -123.0),                 # negativo contabil
        ("(1.234,50)", -1234.5),
        ("R$ 1.500", 1500.0),
        ("-", float("nan")),
        ("", float("nan")),
        ("Receita", float("nan")),
        (None, float("nan")),
    ],
)
def test_conversao_de_numeros(entrada, esperado):
    obtido = para_numero(entrada)
    if np.isnan(esperado):
        assert np.isnan(obtido)
    else:
        assert obtido == pytest.approx(esperado)


@pytest.mark.parametrize(
    "entrada,esperado",
    [
        (2024, 2024),
        ("2024", 2024),
        ("31/12/2024", 2024),
        ("FY2023", 2023),
        ("Exercicio 2022", 2022),
        ("Conta", None),
        (12, None),
    ],
)
def test_extracao_de_anos(entrada, esperado):
    assert extrair_ano(entrada) == esperado


def test_normalizacao_remove_acento_e_pontuacao():
    assert normalizar("Receita de Venda de Bens e/ou Serviços") == (
        "receita de venda de bens e ou servicos"
    )


# ---------------------------------------------------------------------------
# Reconhecimento de contas
# ---------------------------------------------------------------------------


def test_reconhece_por_codigo_cvm():
    resultado = reconhecer("3.01 Receita de Venda de Bens e/ou Serviços")
    assert resultado.chave == "receita_liquida"
    assert resultado.confianca == 1.0


def test_reconhece_por_sinonimo_em_ingles():
    assert reconhecer("Net Revenue").chave == "receita_liquida"
    assert reconhecer("Total Assets").chave == "ativo_total"
    assert reconhecer("EBIT").chave == "ebit"


def test_reconhece_por_sinonimo_em_portugues():
    assert reconhecer("Patrimônio Líquido").chave == "patrimonio_liquido"
    assert reconhecer("Estoques").chave == "estoques"


def test_rotulo_desconhecido_nao_e_forcado():
    resultado = reconhecer("Outras receitas diversas nao operacionais de terceiros")
    assert resultado.chave is None


def test_codigo_cvm_tem_prioridade_sobre_o_texto():
    """O codigo e evidencia mais forte que o nome, que varia entre empresas."""
    resultado = reconhecer("3.05 Resultado Antes do Resultado Financeiro e dos Tributos")
    assert resultado.chave == "ebit"
    assert "codigo CVM" in resultado.motivo


# ---------------------------------------------------------------------------
# Construcao de planilhas de teste nos tres formatos
# ---------------------------------------------------------------------------

ANOS = [2022, 2023, 2024]

DADOS_BASE = {
    "receita_liquida": [1000.0, 1150.0, 1300.0],
    "custo_produtos_vendidos": [600.0, 680.0, 760.0],
    "despesas_operacionais": [220.0, 250.0, 280.0],
    "depreciacao_amortizacao": [50.0, 55.0, 60.0],
    "ebit": [180.0, 220.0, 260.0],
    "resultado_financeiro": [-60.0, -65.0, -70.0],
    "impostos": [40.0, 52.0, 64.0],
    "lucro_liquido": [80.0, 103.0, 126.0],
    "caixa_equivalentes": [150.0, 180.0, 210.0],
    "contas_receber": [200.0, 230.0, 260.0],
    "estoques": [120.0, 140.0, 155.0],
    "imobilizado": [700.0, 760.0, 820.0],
    "ativo_total": [1400.0, 1560.0, 1720.0],
    "fornecedores": [130.0, 150.0, 165.0],
    "divida_curto_prazo": [100.0, 110.0, 120.0],
    "divida_longo_prazo": [400.0, 430.0, 450.0],
    "passivo_total": [1400.0, 1560.0, 1720.0],
    "patrimonio_liquido": [600.0, 690.0, 790.0],
    "capex": [90.0, 100.0, 110.0],
}


def _escrever(caminho, linhas: list[list], aba: str = "Planilha1"):
    pd.DataFrame(linhas).to_excel(caminho, index=False, header=False, sheet_name=aba)
    return caminho


def _planilha_cvm(tmp_path):
    """Export estilo CVM/B3: codigo de conta e custos com sinal negativo."""
    linhas = [
        ["Demonstração do Resultado - Consolidado"],
        ["Conta", "Descrição", *ANOS],
        ["3.01", "Receita de Venda de Bens e/ou Serviços", *DADOS_BASE["receita_liquida"]],
        ["3.02", "Custo dos Bens e/ou Serviços Vendidos", *[-v for v in DADOS_BASE["custo_produtos_vendidos"]]],
        ["3.03", "Resultado Bruto", 400.0, 470.0, 540.0],
        ["3.04", "Despesas/Receitas Operacionais", *[-v for v in DADOS_BASE["despesas_operacionais"]]],
        ["3.05", "Resultado Antes do Resultado Financeiro e dos Tributos", *DADOS_BASE["ebit"]],
        ["3.06", "Resultado Financeiro", *DADOS_BASE["resultado_financeiro"]],
        ["3.08", "Imposto de Renda e Contribuição Social sobre o Lucro", *[-v for v in DADOS_BASE["impostos"]]],
        ["3.11", "Lucro/Prejuízo Consolidado do Período", *DADOS_BASE["lucro_liquido"]],
        [],
        ["Balanço Patrimonial Ativo"],
        ["Conta", "Descrição", *ANOS],
        ["1", "Ativo Total", *DADOS_BASE["ativo_total"]],
        ["1.01.01", "Caixa e Equivalentes de Caixa", *DADOS_BASE["caixa_equivalentes"]],
        ["1.01.03", "Contas a Receber", *DADOS_BASE["contas_receber"]],
        ["1.01.04", "Estoques", *DADOS_BASE["estoques"]],
        ["1.02.03", "Imobilizado", *DADOS_BASE["imobilizado"]],
        ["2", "Passivo Total", *DADOS_BASE["passivo_total"]],
        ["2.01.02", "Fornecedores", *DADOS_BASE["fornecedores"]],
        ["2.01.04", "Empréstimos e Financiamentos", *DADOS_BASE["divida_curto_prazo"]],
        ["2.02.01", "Empréstimos e Financiamentos", *DADOS_BASE["divida_longo_prazo"]],
        ["2.03", "Patrimônio Líquido Consolidado", *DADOS_BASE["patrimonio_liquido"]],
        [],
        ["Demonstração do Fluxo de Caixa"],
        ["Conta", "Descrição", *ANOS],
        ["6.01", "Caixa Líquido Atividades Operacionais", 200.0, 230.0, 260.0],
        ["", "Depreciação e Amortização", *DADOS_BASE["depreciacao_amortizacao"]],
        ["", "Aquisição de Imobilizado", *[-v for v in DADOS_BASE["capex"]]],
    ]
    return _escrever(tmp_path / "cvm.xlsx", linhas)


def _planilha_terminal(tmp_path):
    """Export estilo terminal: nomes em ingles, custos positivos, ruido no topo."""
    linhas = [
        ["Company XYZ SA"],
        ["Currency: BRL", "Units: Thousands"],
        [""],
        ["", "FY2022", "FY2023", "FY2024"],
        ["Net Revenue", *DADOS_BASE["receita_liquida"]],
        ["Cost of Goods Sold", *DADOS_BASE["custo_produtos_vendidos"]],
        ["Gross Profit", 400.0, 470.0, 540.0],
        ["Operating Expenses", *DADOS_BASE["despesas_operacionais"]],
        ["Depreciation and Amortization", *DADOS_BASE["depreciacao_amortizacao"]],
        ["EBIT", *DADOS_BASE["ebit"]],
        ["Net Financial Result", *DADOS_BASE["resultado_financeiro"]],
        ["Income Tax Expense", *DADOS_BASE["impostos"]],
        ["Net Income", *DADOS_BASE["lucro_liquido"]],
        ["Cash and Cash Equivalents", *DADOS_BASE["caixa_equivalentes"]],
        ["Accounts Receivable", *DADOS_BASE["contas_receber"]],
        ["Inventories", *DADOS_BASE["estoques"]],
        ["Property Plant and Equipment", *DADOS_BASE["imobilizado"]],
        ["Total Assets", *DADOS_BASE["ativo_total"]],
        ["Accounts Payable", *DADOS_BASE["fornecedores"]],
        ["Short Term Debt", *DADOS_BASE["divida_curto_prazo"]],
        ["Long Term Debt", *DADOS_BASE["divida_longo_prazo"]],
        ["Total Liabilities and Equity", *DADOS_BASE["passivo_total"]],
        ["Total Shareholders Equity", *DADOS_BASE["patrimonio_liquido"]],
        ["Capital Expenditures", *[-v for v in DADOS_BASE["capex"]]],
    ]
    return _escrever(tmp_path / "terminal.xlsx", linhas)


@pytest.fixture
def cvm(tmp_path):
    return _planilha_cvm(tmp_path)


@pytest.fixture
def terminal(tmp_path):
    return _planilha_terminal(tmp_path)


# ---------------------------------------------------------------------------
# Importacao ponta a ponta
# ---------------------------------------------------------------------------


def test_importa_export_cvm(cvm):
    dfs = importar(cvm, empresa="Teste S.A.")
    assert dfs.anos == ANOS
    assert dfs.serie("receita_liquida").tolist() == DADOS_BASE["receita_liquida"]
    assert dfs.serie("ebit").tolist() == DADOS_BASE["ebit"]
    assert dfs.serie("patrimonio_liquido").tolist() == DADOS_BASE["patrimonio_liquido"]


def test_importa_export_terminal(terminal):
    dfs = importar(terminal, empresa="Teste S.A.")
    assert dfs.anos == ANOS
    assert dfs.serie("receita_liquida").tolist() == DADOS_BASE["receita_liquida"]
    assert dfs.serie("ebit").tolist() == DADOS_BASE["ebit"]


def test_as_duas_origens_chegam_ao_mesmo_resultado(cvm, terminal):
    """O ponto do vocabulario canonico: origem diferente, numero identico."""
    a = importar(cvm).valores
    b = importar(terminal).valores
    comuns = sorted(set(a.index) & set(b.index))
    assert len(comuns) >= 15
    pd.testing.assert_frame_equal(
        a.loc[comuns].astype(float), b.loc[comuns].astype(float)
    )


def test_sinal_de_custos_e_padronizado(cvm, terminal):
    """CVM traz custo negativo e o terminal positivo; ambos viram magnitude."""
    for caminho in (cvm, terminal):
        dfs = importar(caminho)
        assert (dfs.serie("custo_produtos_vendidos").dropna() > 0).all()
        assert (dfs.serie("impostos").dropna() > 0).all()
        assert (dfs.serie("capex").dropna() > 0).all()


def test_resultado_financeiro_mantem_o_sinal(cvm):
    """Resultado financeiro pode ser negativo de verdade e nao pode virar magnitude."""
    dfs = importar(cvm)
    assert (dfs.serie("resultado_financeiro").dropna() < 0).all()


def test_derivacoes_preenchem_o_que_falta(cvm):
    dfs = importar(cvm)
    # LAIR nao existe no arquivo: sai de EBIT + resultado financeiro.
    assert "lucro_antes_impostos" in dfs.derivadas
    assert dfs.serie("lucro_antes_impostos").tolist() == [120.0, 155.0, 190.0]


def test_metricas_derivadas(cvm):
    dfs = importar(cvm)
    assert dfs.ebitda().tolist() == [230.0, 275.0, 320.0]
    assert dfs.divida_bruta().tolist() == [500.0, 540.0, 570.0]
    assert dfs.divida_liquida().tolist() == [350.0, 360.0, 360.0]
    # Capital de giro = recebiveis + estoques - fornecedores
    assert dfs.capital_giro().tolist() == [190.0, 220.0, 250.0]


def test_ano_base_e_o_mais_recente(cvm):
    assert importar(cvm).ano_base == 2024


def test_limite_de_anos(cvm):
    dfs = importar(cvm, anos_maximos=2)
    assert dfs.anos == [2023, 2024]


def test_linhas_nao_reconhecidas_sao_reportadas(tmp_path):
    """Nada e descartado em silencio: o que nao foi entendido volta para o usuario."""
    linhas = [
        ["Conta", 2023, 2024],
        ["Receita Líquida", 100.0, 110.0],
        ["EBIT", 20.0, 22.0],
        ["Lucro Líquido", 10.0, 11.0],
        ["Ativo Total", 200.0, 220.0],
        ["Patrimônio Líquido", 90.0, 100.0],
        ["Caixa e Equivalentes de Caixa", 30.0, 35.0],
        ["Rubrica exotica da empresa", 5.0, 6.0],
    ]
    dfs = importar(_escrever(tmp_path / "x.xlsx", linhas))
    rotulos = [linha.rotulo for linha in dfs.nao_reconhecidas]
    assert "Rubrica exotica da empresa" in rotulos


def test_conta_obrigatoria_ausente_vira_aviso(tmp_path):
    linhas = [
        ["Conta", 2023, 2024],
        ["Receita Líquida", 100.0, 110.0],
        ["EBIT", 20.0, 22.0],
    ]
    dfs = importar(_escrever(tmp_path / "x.xlsx", linhas))
    assert any("obrigatorias" in aviso for aviso in dfs.avisos)


def test_balanco_que_nao_fecha_vira_aviso(tmp_path):
    linhas = [
        ["Conta", 2023, 2024],
        ["Receita Líquida", 100.0, 110.0],
        ["EBIT", 20.0, 22.0],
        ["Lucro Líquido", 10.0, 11.0],
        ["Ativo Total", 200.0, 220.0],
        ["Passivo Total", 150.0, 160.0],  # nao bate com o ativo
        ["Patrimônio Líquido", 90.0, 100.0],
        ["Caixa e Equivalentes de Caixa", 30.0, 35.0],
    ]
    dfs = importar(_escrever(tmp_path / "x.xlsx", linhas))
    assert any("divergem" in aviso for aviso in dfs.avisos)


def test_planilha_sem_anos_e_rejeitada(tmp_path):
    linhas = [["Conta", "Valor"], ["Receita", 100.0]]
    with pytest.raises(ValueError, match="nenhuma tabela com anos"):
        importar(_escrever(tmp_path / "x.xlsx", linhas))


def test_arquivo_inexistente(tmp_path):
    with pytest.raises(FileNotFoundError):
        importar(tmp_path / "nao_existe.xlsx")


def test_escalar_converte_unidade(cvm):
    dfs = importar(cvm, unidade="R$ mil").escalar(1000, "R$ milhoes")
    assert dfs.serie("receita_liquida").tolist() == [1.0, 1.15, 1.3]
    assert dfs.unidade == "R$ milhoes"


def test_aviso_de_escala_usa_milhar_com_ponto(cvm):
    """O aviso vai direto para a tela: "1,000,000" se le errado em portugues."""
    dfs = importar(cvm).escalar(1_000_000, "R$ milhoes")
    assert "Valores divididos por 1.000.000." in dfs.avisos


def test_mapeamento_manual_corrige_reconhecimento(tmp_path):
    linhas = [
        ["Conta", 2023, 2024],
        ["Receita Líquida", 100.0, 110.0],
        ["EBIT", 20.0, 22.0],
        ["Lucro Líquido", 10.0, 11.0],
        ["Ativo Total", 200.0, 220.0],
        ["Patrimônio Líquido", 90.0, 100.0],
        ["Caixa e Equivalentes de Caixa", 30.0, 35.0],
        ["Verba de reposicao de ativos", 7.0, 8.0],
    ]
    caminho = _escrever(tmp_path / "x.xlsx", linhas)
    dfs = importar(caminho)
    assert not dfs.tem("capex")

    corrigido = aplicar_mapeamento_manual(
        dfs, caminho, {"Verba de reposicao de ativos": "capex"}
    )
    assert corrigido.serie("capex").tolist() == [7.0, 8.0]
    assert "Verba de reposicao de ativos" not in [
        linha.rotulo for linha in corrigido.nao_reconhecidas
    ]


def test_mapeamento_para_conta_inexistente_e_rejeitado(cvm):
    dfs = importar(cvm)
    with pytest.raises(ValueError, match="inexistentes"):
        aplicar_mapeamento_manual(dfs, cvm, {"Qualquer": "conta_que_nao_existe"})


# ---------------------------------------------------------------------------
# Template
# ---------------------------------------------------------------------------


def test_template_e_gerado_e_reimportavel(tmp_path):
    """O template precisa voltar pelo mesmo caminho das demais origens."""
    from openpyxl import load_workbook

    caminho = gerar_template(tmp_path / "template.xlsx", anos=[2023, 2024])
    assert caminho.exists()

    wb = load_workbook(caminho)
    assert "Instrucoes" in wb.sheetnames
    ws = wb["Demonstracoes"]

    # Preenche as contas obrigatorias como um usuario faria, usando os rotulos
    # exatamente como o template os escreve.
    valores = {
        "Receita liquida": (100.0, 110.0),
        "EBIT (resultado operacional)": (20.0, 22.0),
        "Lucro liquido": (10.0, 11.0),
        "Ativo total": (200.0, 220.0),
        "Patrimonio liquido": (90.0, 100.0),
        "Caixa e equivalentes de caixa": (30.0, 35.0),
    }
    for linha in range(1, ws.max_row + 1):
        rotulo = ws.cell(row=linha, column=1).value
        if rotulo in valores:
            ws.cell(row=linha, column=2, value=valores[rotulo][0])
            ws.cell(row=linha, column=3, value=valores[rotulo][1])
    wb.save(caminho)

    dfs = importar(caminho, empresa="Template S.A.")
    assert dfs.anos == [2023, 2024]
    assert dfs.serie("receita_liquida").tolist() == [100.0, 110.0]
    assert dfs.serie("patrimonio_liquido").tolist() == [90.0, 100.0]
    assert dfs.serie("ebit").tolist() == [20.0, 22.0]
    assert not any("obrigatorias" in aviso for aviso in dfs.avisos)


def test_grade_e_localizada_com_ruido_no_topo(terminal):
    from valuation.importacao import carregar_abas

    abas = carregar_abas(terminal)
    grade = localizar_grade(next(iter(abas.values())))
    assert grade is not None
    assert grade.anos == ANOS
    assert grade.coluna_rotulos == 0


# ---------------------------------------------------------------------------
# Demonstracao de holding: sem receita operacional
# ---------------------------------------------------------------------------


def _conferido(**contas) -> list[str]:
    """Roda ``_conferir`` sobre um quadro montado a mão e devolve os avisos."""
    import pandas as pd

    from valuation.importacao.importador import _conferir

    avisos: list[str] = []
    _conferir(pd.DataFrame({2024: contas}), avisos)
    return avisos


def test_demonstracao_sem_receita_com_lucro_e_de_holding():
    """Margem, giro e capex sobre receita não descrevem nada aqui.

    O resultado vem de equivalência patrimonial das investidas. Medido no
    consolidado de 2024, o corte em 10% do lucro pega 11 das 467 companhias —
    entre elas BB Seguridade (lucro de R$ 8,7 bi **sem nenhuma linha de
    receita**) e Caixa Seguridade.
    """
    avisos = _conferido(receita_liquida=0.0, lucro_liquido=3_765_000_000.0)
    assert any("holding" in a for a in avisos), avisos


def test_receita_ausente_conta_como_zero():
    """Não é "não dá para saber": é o caso mais forte do mesmo sinal.

    A BB Seguridade dá lucro de R$ 8,7 bi sem nenhuma linha de receita
    reconhecida, e tratar isso como dado faltante deixaria passar justamente a
    companhia em que o aviso mais importa.
    """
    avisos = _conferido(lucro_liquido=8_703_000_000.0)
    assert any("holding" in a for a in avisos), avisos


def test_operacao_de_verdade_nao_dispara_o_sinal():
    """Sinal que dispara em quem opera não dirige atenção, gasta.

    A Itaúsa tem lucro maior que a receita e **não** entra: o consolidado dela
    tem R$ 8,2 bi de receita de verdade, das controladas operacionais. Ela é
    coberta pelo sinal de item não recorrente, que é outro assunto.
    """
    assert not any(
        "holding" in a
        for a in _conferido(
            receita_liquida=8_235_000_000.0, lucro_liquido=14_887_000_000.0
        )
    )
    assert not any(
        "holding" in a
        for a in _conferido(
            receita_liquida=37_986_941_000.0, lucro_liquido=6_318_800_000.0
        )
    )


def test_sem_lucro_nao_ha_o_que_afirmar():
    """Companhia sem receita e sem lucro não é holding — é papel em branco."""
    assert not any("holding" in a for a in _conferido(receita_liquida=0.0, lucro_liquido=0.0))


# ---------------------------------------------------------------------------
# Ativo e passivo nao se confundem
# ---------------------------------------------------------------------------


def test_o_rotulo_nao_move_uma_linha_de_ativo_para_conta_de_passivo():
    """O rótulo tem prioridade sobre o código — mas não atravessa o balanço.

    Foi a prioridade do rótulo que corrigiu o patrimônio líquido dos bancos,
    onde `2.07` quer dizer outra coisa. O limite que faltava: no BPA, `1.03.02`
    se chama "Imposto de Renda e Contribuição Social - Diferidos", **o mesmo
    rótulo** do passivo `2.02.03`, e o casamento por rótulo punha o ativo fiscal
    diferido na conta de passivo.

    Medido no DFP consolidado de 2025: **12 companhias, R$ 275,0 bilhões** — o
    Bradesco com R$ 111,2 bi, o Banco do Brasil com R$ 89,3 bi. Todas bancos,
    porque `1.03` só existe no plano financeiro. Nenhuma identidade denunciava:
    os dois lados continuam fechando, porque a linha some de um total e aparece
    noutro que ninguém soma de volta.
    """
    from valuation.importacao.esquema import reconhecer

    rotulo = "Imposto de Renda e Contribuição Social - Diferidos"

    # Sem saber o lado, o rótulo casa com a conta de passivo — é o caso da
    # planilha colada à mão, onde não há código.
    assert reconhecer(rotulo, demonstracao="bp").chave == "tributos_diferidos_passivo"

    # Sabendo que a linha está no ativo, a conta de passivo é recusada.
    no_ativo = reconhecer(rotulo, demonstracao="bp", codigo_da_linha="1.03.02")
    assert no_ativo.chave != "tributos_diferidos_passivo"

    # E o lado certo continua passando.
    no_passivo = reconhecer(rotulo, demonstracao="bp", codigo_da_linha="2.02.03")
    assert no_passivo.chave == "tributos_diferidos_passivo"


def test_a_guarda_de_lado_so_recusa_quando_os_dois_lados_sao_conhecidos():
    """Ausência de evidência não é evidência de erro.

    A guarda tem de ser inerte fora do balanço e quando o código não diz o lado
    — senão ela derrubaria reconhecimento legítimo, que é pior que o defeito:
    conta vazia some da tela sem explicação. Medido na base de 2025, a cobertura
    não se moveu (mediana de 82 contas por companhia antes e depois); só o
    mínimo caiu de 32 para 31, que é a linha corretamente recusada.
    """
    from valuation.importacao.esquema import _lado_confere

    # Conta de DRE: a raiz `3` nao separa lados, e `demonstracao=` ja resolve.
    assert _lado_confere("receita_liquida", "3.01")
    # Sem codigo nao ha o que conferir.
    assert _lado_confere("tributos_diferidos_passivo", None)
    assert _lado_confere("tributos_diferidos_passivo", "")
    # Conta que nao declara lado unico tambem passa.
    assert _lado_confere(None, "1.03.02")


def test_operacoes_continuadas_nao_sao_o_lucro_liquido():
    """`3.09` deixou de ser código alternativo de `lucro_liquido`.

    Ele é "Resultado Líquido das Operações Continuadas" e não o consolidado do
    período: onde há operação descontinuada, os dois são números diferentes.
    Estava declarado como alternativa para quando `3.11` falta — e no ITR ele
    **ganhava**, em 195 companhias.

    Medido no ITR de 2026: em 186 delas `3.10` é zero e tanto faz, mas em **9
    não**. Na Randoncorp com o **sinal trocado** — o app lia +R$ 56,5 mi de
    lucro onde a companhia publicou −R$ 23,1 mi de prejuízo; na CBD, −203,0
    contra −252,0. Depois da correção: 5 companhias caem em `3.09`, todas sem
    `3.11` publicado, e **nenhuma diverge do consolidado**.
    """
    from valuation.importacao.esquema import POR_CHAVE

    assert POR_CHAVE["lucro_liquido"].codigos_cvm == ("3.11",)
    assert POR_CHAVE["resultado_continuadas"].codigos_cvm == ("3.09",)


def test_o_lucro_liquido_ausente_vem_da_soma_das_duas_pontas(tmp_path):
    """O recurso continua existindo, e agora pela identidade.

    Somar continuadas e descontinuadas **é** o consolidado; pegar uma das duas é
    outra conta. É a mesma identidade que a DRE gerencial já verifica.
    """
    caminho = _escrever(
        tmp_path / "continuadas.xlsx",
        [
            ["Demonstração do Resultado - Consolidado"],
            ["Conta", "Descrição", 2024],
            ["3.01", "Receita de Venda de Bens e/ou Serviços", 1000.0],
            ["3.05", "Resultado Antes do Resultado Financeiro e dos Tributos", 200.0],
            ["3.09", "Resultado Líquido das Operações Continuadas", 100.0],
            ["3.10", "Resultado Líquido de Operações Descontinuadas", -30.0],
        ],
    )
    dfs = importar(caminho)

    assert float(dfs.valor("resultado_continuadas")) == pytest.approx(100.0)
    assert float(dfs.valor("operacoes_descontinuadas")) == pytest.approx(-30.0)
    # 100 - 30 = 70, e nao os 100 que a leitura antiga devolveria.
    assert float(dfs.valor("lucro_liquido")) == pytest.approx(70.0)

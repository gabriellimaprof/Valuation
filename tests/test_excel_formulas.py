"""Verifica que as formulas gravadas no Excel reproduzem os numeros do Python.

Este e o teste mais importante do exportador: uma planilha bonita que calcula
diferente do motor e pior do que nenhuma planilha, porque o erro so aparece na
mao de quem recebeu o arquivo.

O teste depende do pacote ``formulas`` para avaliar o workbook fora do Excel;
sem ele, o teste e pulado em vez de dar falso positivo.
"""

from __future__ import annotations

import pytest

from valuation import avaliar, exportar_excel

formulas = pytest.importorskip("formulas", reason="pacote 'formulas' nao instalado")


def _valores_calculados(caminho) -> dict[str, float]:
    """Avalia o workbook e devolve o mapa de celula -> valor calculado."""
    modelo = formulas.ExcelModel().loads(str(caminho)).finish()
    solucao = modelo.calculate()
    valores = {}
    for chave, valor in solucao.items():
        try:
            valores[chave.upper()] = float(valor.value[0, 0])
        except (AttributeError, TypeError, ValueError, IndexError):
            continue
    return valores


def _buscar(valores: dict[str, float], aba: str, celula: str, caminho) -> float:
    """Localiza uma celula no resultado, que vem chaveado pelo caminho do arquivo."""
    alvo = f"'[{caminho.name.upper()}]{aba.upper()}'!{celula.upper()}"
    if alvo in valores:
        return valores[alvo]
    sufixo = f"]{aba.upper()}'!{celula.upper()}"
    candidatos = [v for k, v in valores.items() if k.endswith(sufixo)]
    if not candidatos:
        raise AssertionError(f"celula {aba}!{celula} nao encontrada no workbook avaliado")
    return candidatos[0]


def _localizar_linha(ws, rotulo: str) -> int:
    for linha in range(1, ws.max_row + 1):
        if ws.cell(row=linha, column=1).value == rotulo:
            return linha
    raise AssertionError(f"linha {rotulo!r} nao encontrada em {ws.title}")


@pytest.mark.parametrize("meio_de_ano", [False, True])
def test_excel_reproduz_o_motor_python(empresa_exemplo, tmp_path, meio_de_ano):
    from openpyxl import load_workbook

    resultado = avaliar(empresa_exemplo, meio_de_ano=meio_de_ano)
    caminho = tmp_path / "modelo.xlsx"
    exportar_excel(resultado, caminho)

    valores = _valores_calculados(caminho)
    wb = load_workbook(caminho)

    linha_wacc = _localizar_linha(wb["Custo de Capital"], "WACC (BRL nominal)")
    wacc_excel = _buscar(valores, "Custo de Capital", f"B{linha_wacc}", caminho)
    assert wacc_excel == pytest.approx(resultado.custo_capital.wacc_brl, rel=1e-9)

    aba_dcf = wb["DCF"]
    for rotulo, esperado in [
        ("VP dos fluxos do periodo explicito", resultado.dcf.valor_presente_explicito),
        ("Valor terminal (fim do ano n)", resultado.dcf.valor_terminal),
        ("VP do valor terminal", resultado.dcf.valor_presente_terminal),
        ("Enterprise Value", resultado.dcf.enterprise_value),
        ("Equity Value", resultado.dcf.equity_value),
    ]:
        linha = _localizar_linha(aba_dcf, rotulo)
        obtido = _buscar(valores, "DCF", f"B{linha}", caminho)
        assert obtido == pytest.approx(esperado, rel=1e-9), rotulo


def test_excel_reproduz_a_projecao(empresa_exemplo, tmp_path):
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    resultado = avaliar(empresa_exemplo)
    caminho = tmp_path / "projecao.xlsx"
    exportar_excel(resultado, caminho)

    valores = _valores_calculados(caminho)
    ws = load_workbook(caminho)["Projecao"]
    n = resultado.projecao.horizonte

    for rotulo, esperado in [
        ("Receita liquida", resultado.projecao.receita),
        ("EBITDA", resultado.projecao.ebitda),
        ("NOPAT", resultado.projecao.nopat),
        ("FCFF (fluxo para a firma)", resultado.projecao.fcff),
    ]:
        linha = _localizar_linha(ws, rotulo)
        for i in range(n):
            celula = f"{get_column_letter(3 + i)}{linha}"
            obtido = _buscar(valores, "Projecao", celula, caminho)
            assert obtido == pytest.approx(float(esperado[i]), rel=1e-9), (
                f"{rotulo} ano {i + 1}"
            )


def test_excel_reproduz_multiplo_de_saida(empresa_exemplo, tmp_path):
    """A perpetuidade por multiplo usa um ramo de formula diferente do Gordon."""
    from openpyxl import load_workbook

    from valuation import substituir_varios

    empresa = substituir_varios(
        empresa_exemplo,
        {
            "perpetuidade.metodo": "multiplo",
            "perpetuidade.multiplo_saida": 7.5,
        },
    )
    resultado = avaliar(empresa)
    caminho = tmp_path / "multiplo.xlsx"
    exportar_excel(resultado, caminho)

    valores = _valores_calculados(caminho)
    ws = load_workbook(caminho)["DCF"]
    linha = _localizar_linha(ws, "Valor terminal (fim do ano n)")
    obtido = _buscar(valores, "DCF", f"B{linha}", caminho)
    assert obtido == pytest.approx(resultado.dcf.valor_terminal, rel=1e-9)

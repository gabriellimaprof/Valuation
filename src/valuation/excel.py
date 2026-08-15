"""Exportacao do modelo para Excel com formulas vivas.

As abas Premissas, Custo de Capital, Projecao e DCF sao escritas com **formulas
do Excel**, nao com valores calculados no Python. Quem receber o arquivo pode
alterar uma premissa e ver o valuation inteiro recalcular, e um revisor
consegue rastrear cada numero ate sua origem -- que e exatamente o que se
espera de um modelo de valuation entregue a terceiros.

As abas de Multiplos, Sensibilidade e Monte Carlo trazem valores calculados no
Python (uma tabela de sensibilidade viva exigiria replicar o modelo inteiro por
celula), e vem rotuladas como tal.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .modelo import ResultadoValuation
from .multiplos import Alvo, Comparavel, avaliar_por_multiplos, estatisticas, tabela_comparaveis
from .sensibilidade import ResultadoSimulacao

PCT = "0.00%"
PCT1 = "0.0%"
MOEDA = "#,##0.0"
NUMERO = "0.00"
MULTIPLO = '0.00"x"'

AZUL_ENTRADA = Font(color="0000CC", bold=True)
PRETO_FORMULA = Font(color="000000")
VERDE_LIGACAO = Font(color="007700")

FUNDO_TITULO = PatternFill("solid", fgColor="1F3864")
FUNDO_SECAO = PatternFill("solid", fgColor="D9E1F2")
FUNDO_TOTAL = PatternFill("solid", fgColor="FFF2CC")

BORDA_TOPO = Border(top=Side(style="thin", color="808080"))


@dataclass
class _Aba:
    """Escritor sequencial de uma aba, que devolve referencias absolutas.

    Guardar a referencia de cada celula escrita e o que permite montar as
    formulas das abas seguintes sem endereços fixos espalhados pelo codigo.
    """

    ws: Worksheet
    linha: int = 1

    def _ref(self, coluna: str, linha: int) -> str:
        return f"'{self.ws.title}'!${coluna}${linha}"

    def titulo(self, texto: str, largura: int = 8) -> None:
        celula = self.ws.cell(row=self.linha, column=1, value=texto)
        celula.font = Font(bold=True, size=14, color="FFFFFF")
        celula.fill = FUNDO_TITULO
        for coluna in range(2, largura + 1):
            self.ws.cell(row=self.linha, column=coluna).fill = FUNDO_TITULO
        self.linha += 2

    def secao(self, texto: str, largura: int = 8) -> None:
        celula = self.ws.cell(row=self.linha, column=1, value=texto)
        celula.font = Font(bold=True)
        for coluna in range(1, largura + 1):
            self.ws.cell(row=self.linha, column=coluna).fill = FUNDO_SECAO
        self.linha += 1

    def entrada(self, rotulo: str, valor, formato: str = NUMERO) -> str:
        """Premissa editavel pelo usuario (azul, na convencao de modelagem)."""
        return self._campo(rotulo, valor, formato, AZUL_ENTRADA)

    def formula(self, rotulo: str, expressao: str, formato: str = NUMERO) -> str:
        """Celula calculada dentro da propria aba (preta)."""
        return self._campo(rotulo, expressao, formato, PRETO_FORMULA)

    def ligacao(self, rotulo: str, expressao: str, formato: str = NUMERO) -> str:
        """Celula que traz um numero de outra aba (verde)."""
        return self._campo(rotulo, expressao, formato, VERDE_LIGACAO)

    def _campo(self, rotulo: str, valor, formato: str, fonte: Font) -> str:
        self.ws.cell(row=self.linha, column=1, value=rotulo)
        celula = self.ws.cell(row=self.linha, column=2, value=valor)
        celula.number_format = formato
        celula.font = fonte
        ref = self._ref("B", self.linha)
        self.linha += 1
        return ref

    def total(self, rotulo: str, expressao: str, formato: str = MOEDA) -> str:
        self.ws.cell(row=self.linha, column=1, value=rotulo).font = Font(bold=True)
        celula = self.ws.cell(row=self.linha, column=2, value=expressao)
        celula.number_format = formato
        celula.font = Font(bold=True)
        celula.fill = FUNDO_TOTAL
        celula.border = BORDA_TOPO
        ref = self._ref("B", self.linha)
        self.linha += 1
        return ref

    def nota(self, texto: str) -> None:
        celula = self.ws.cell(row=self.linha, column=1, value=texto)
        celula.font = Font(italic=True, size=9, color="666666")
        self.linha += 1

    def pular(self, n: int = 1) -> None:
        self.linha += n

    def cabecalho_anos(self, anos: list[int], coluna_base: bool = True) -> int:
        """Escreve a linha de cabecalho da tabela anual e devolve sua linha."""
        self.ws.cell(row=self.linha, column=1, value="").font = Font(bold=True)
        if coluna_base:
            celula = self.ws.cell(row=self.linha, column=2, value="Base")
            celula.font = Font(bold=True)
            celula.alignment = Alignment(horizontal="center")
        for i, ano in enumerate(anos):
            celula = self.ws.cell(row=self.linha, column=3 + i, value=f"Ano {ano}")
            celula.font = Font(bold=True)
            celula.alignment = Alignment(horizontal="center")
            celula.fill = FUNDO_SECAO
        linha = self.linha
        self.linha += 1
        return linha

    def linha_anual(
        self,
        rotulo: str,
        valores: list,
        formato: str = MOEDA,
        base=None,
        fonte: Font | None = None,
        negrito: bool = False,
    ) -> dict[str, str]:
        """Escreve uma linha da tabela anual e devolve as referencias por coluna.

        A chave ``"base"`` aponta para a coluna B (ano 0) quando ``base`` e
        informado; as demais chaves sao os indices ``0..n-1`` dos anos.
        """
        self.ws.cell(row=self.linha, column=1, value=rotulo).font = Font(bold=negrito)
        refs: dict[str, str] = {}
        fonte = fonte or (Font(bold=True) if negrito else None)

        if base is not None:
            celula = self.ws.cell(row=self.linha, column=2, value=base)
            celula.number_format = formato
            if fonte:
                celula.font = fonte
            refs["base"] = self._ref("B", self.linha)

        for i, valor in enumerate(valores):
            coluna = 3 + i
            celula = self.ws.cell(row=self.linha, column=coluna, value=valor)
            celula.number_format = formato
            if fonte:
                celula.font = fonte
            if negrito:
                celula.fill = FUNDO_TOTAL
                celula.border = BORDA_TOPO
            refs[str(i)] = self._ref(get_column_letter(coluna), self.linha)
        self.linha += 1
        return refs

    def ajustar(self, largura_rotulo: int = 38, largura_dados: int = 14, n: int = 12) -> None:
        self.ws.column_dimensions["A"].width = largura_rotulo
        for coluna in range(2, n + 3):
            self.ws.column_dimensions[get_column_letter(coluna)].width = largura_dados


def _celula(ref: str) -> str:
    """Extrai o endereco puro (``$B$9``) de uma referencia com nome de aba."""
    return ref.split("!")[-1]


def exportar_excel(
    resultado: ResultadoValuation,
    caminho: str | Path,
    sensibilidade: pd.DataFrame | None = None,
    comparaveis: list[Comparavel] | None = None,
    alvo: Alvo | None = None,
    simulacao: ResultadoSimulacao | None = None,
    cenarios: pd.DataFrame | None = None,
    retorno=None,
    acionista=None,
) -> Path:
    """Gera a planilha completa do valuation e devolve o caminho escrito."""
    caminho = Path(caminho)
    wb = Workbook()
    wb.remove(wb.active)

    refs_premissas = _aba_premissas(wb, resultado)
    refs_cc = _aba_custo_capital(wb, resultado, refs_premissas)
    refs_proj = _aba_projecao(wb, resultado, refs_premissas)
    _aba_dcf(wb, resultado, refs_premissas, refs_cc, refs_proj)

    if retorno is not None:
        _aba_retorno(wb, resultado, retorno, acionista)

    if comparaveis:
        _aba_multiplos(wb, comparaveis, alvo)
    if sensibilidade is not None:
        _aba_sensibilidade(wb, sensibilidade)
    if cenarios is not None:
        _aba_cenarios(wb, cenarios)
    if simulacao is not None:
        _aba_monte_carlo(wb, simulacao)

    _aba_leia_me(wb, resultado)

    caminho.parent.mkdir(parents=True, exist_ok=True)
    wb.save(caminho)
    return caminho


def _aba_premissas(wb: Workbook, resultado: ResultadoValuation) -> dict[str, object]:
    empresa = resultado.empresa
    aba = _Aba(wb.create_sheet("Premissas"))
    n = resultado.projecao.horizonte

    aba.titulo(f"{empresa.nome} - Premissas", largura=n + 3)
    aba.nota(
        f"Data-base: {empresa.data_base or 'nao informada'} | "
        f"Moeda: {empresa.moeda} | Unidade: {empresa.unidade}"
    )
    aba.nota("Celulas azuis sao premissas editaveis. Altere e o modelo recalcula.")
    aba.pular()

    refs: dict[str, object] = {}

    aba.secao("Macroeconomicas", largura=n + 3)
    refs["inflacao_brl"] = aba.entrada("Inflacao BRL de longo prazo", empresa.macro.inflacao_brl, PCT)
    refs["inflacao_usd"] = aba.entrada("Inflacao USD de longo prazo", empresa.macro.inflacao_usd, PCT)
    refs["aliquota_ir"] = aba.entrada("Aliquota de IR/CSLL", empresa.macro.aliquota_ir, PCT)
    aba.pular()

    cc = empresa.custo_capital
    aba.secao("Custo de capital", largura=n + 3)
    refs["rf_usd"] = aba.entrada("Taxa livre de risco (USD)", cc.rf_usd, PCT)
    refs["erp_maduro"] = aba.entrada("Premio de risco de mercado maduro", cc.erp_maduro, PCT)
    refs["risco_pais"] = aba.entrada("Premio de risco-pais (Brasil)", cc.risco_pais, PCT)
    refs["lambda_pais"] = aba.entrada("Lambda (exposicao ao risco-pais)", cc.lambda_pais, NUMERO)
    refs["premio_tamanho"] = aba.entrada("Premio de tamanho", cc.premio_tamanho, PCT)
    if cc.beta_desalavancado is not None:
        refs["beta_desalavancado_entrada"] = aba.entrada(
            "Beta desalavancado (informado)", cc.beta_desalavancado, NUMERO
        )
    else:
        refs["beta_setor"] = aba.entrada(
            "Beta alavancado dos comparaveis", cc.beta_alavancado_setor, NUMERO
        )
        refs["divida_pl_setor"] = aba.entrada("D/E dos comparaveis", cc.divida_pl_setor, NUMERO)
    refs["divida_pl_alvo"] = aba.entrada("D/E alvo da empresa", cc.divida_pl_alvo, NUMERO)
    if cc.custo_divida_brl is not None:
        refs["custo_divida_brl"] = aba.entrada(
            "Custo da divida bruto (BRL)", cc.custo_divida_brl, PCT
        )
    else:
        refs["spread_credito"] = aba.entrada("Spread de credito", cc.spread_credito, PCT)
    aba.pular()

    perp = empresa.perpetuidade
    aba.secao("Perpetuidade", largura=n + 3)
    refs["metodo_perp"] = aba.entrada("Metodo", perp.metodo, "General")
    refs["g_perpetuo"] = aba.entrada("Crescimento perpetuo", perp.crescimento_perpetuo, PCT)
    refs["roic_perp"] = aba.entrada(
        "ROIC na perpetuidade (vazio = sem normalizacao)",
        perp.roic_perpetuidade if perp.roic_perpetuidade is not None else "",
        PCT,
    )
    refs["multiplo_saida"] = aba.entrada(
        "Multiplo de saida EV/EBITDA",
        perp.multiplo_saida if perp.multiplo_saida is not None else "",
        MULTIPLO,
    )
    aba.pular()

    ponte = empresa.ponte
    aba.secao("Ponte EV -> Equity (data-base)", largura=n + 3)
    refs["divida_bruta"] = aba.entrada("Divida bruta", ponte.divida_bruta, MOEDA)
    refs["caixa"] = aba.entrada("Caixa e equivalentes", ponte.caixa, MOEDA)
    refs["aplicacoes"] = aba.entrada("Aplicacoes financeiras", ponte.aplicacoes_financeiras, MOEDA)
    refs["minoritarios"] = aba.entrada("Participacao de minoritarios", ponte.minoritarios, MOEDA)
    refs["contingencias"] = aba.entrada("Contingencias", ponte.contingencias, MOEDA)
    refs["deficit_atuarial"] = aba.entrada("Deficit atuarial", ponte.deficit_atuarial, MOEDA)
    refs["ativos_nao_op"] = aba.entrada("Ativos nao operacionais", ponte.ativos_nao_operacionais, MOEDA)
    refs["acoes"] = aba.entrada(
        "Acoes em circulacao",
        ponte.acoes_em_circulacao if ponte.acoes_em_circulacao is not None else "",
        MOEDA,
    )
    aba.pular()

    op = empresa.operacionais
    aba.secao("Operacionais", largura=n + 3)
    refs["receita_base"] = aba.entrada("Receita liquida do ano base", op.receita_base, MOEDA)
    aba.pular()
    aba.cabecalho_anos(resultado.projecao.anos)
    refs["crescimento"] = aba.linha_anual(
        "Crescimento da receita", list(op.crescimento_receita), PCT, fonte=AZUL_ENTRADA
    )
    refs["margem_ebitda"] = aba.linha_anual(
        "Margem EBITDA", list(op.margem_ebitda), PCT, fonte=AZUL_ENTRADA
    )
    refs["dep_pct"] = aba.linha_anual(
        "Depreciacao / Receita", list(op.depreciacao_pct_receita), PCT, fonte=AZUL_ENTRADA
    )
    refs["capex_pct"] = aba.linha_anual(
        "Capex / Receita", list(op.capex_pct_receita), PCT, fonte=AZUL_ENTRADA
    )
    refs["giro_pct"] = aba.linha_anual(
        "Capital de giro / Receita", list(op.capital_giro_pct_receita), PCT, fonte=AZUL_ENTRADA
    )

    aba.ajustar(n=n)
    return refs


def _aba_custo_capital(
    wb: Workbook, resultado: ResultadoValuation, p: dict
) -> dict[str, str]:
    aba = _Aba(wb.create_sheet("Custo de Capital"))
    cc = resultado.empresa.custo_capital

    aba.titulo("Custo de capital (CAPM + risco-pais)")
    aba.nota(
        "Ke montado em USD nominal e convertido para BRL por diferencial de inflacao, "
        "evitando dupla contagem de risco soberano."
    )
    aba.pular()

    refs: dict[str, str] = {}

    aba.secao("Beta")
    if "beta_desalavancado_entrada" in p:
        refs["beta_u"] = aba.ligacao(
            "Beta desalavancado", f"={p['beta_desalavancado_entrada']}", NUMERO
        )
    else:
        refs["beta_u"] = aba.formula(
            "Beta desalavancado (Hamada)",
            f"={p['beta_setor']}/(1+(1-{p['aliquota_ir']})*{p['divida_pl_setor']})",
            NUMERO,
        )
    refs["beta_l"] = aba.formula(
        "Beta realavancado (D/E alvo)",
        f"={refs['beta_u']}*(1+(1-{p['aliquota_ir']})*{p['divida_pl_alvo']})",
        NUMERO,
    )
    aba.pular()

    aba.secao("Custo do capital proprio (Ke)")
    refs["ke_usd"] = aba.formula(
        "Ke em USD nominal",
        f"={p['rf_usd']}+{refs['beta_l']}*{p['erp_maduro']}"
        f"+{p['lambda_pais']}*{p['risco_pais']}+{p['premio_tamanho']}",
        PCT,
    )
    refs["ke_brl"] = aba.formula(
        "Ke em BRL nominal",
        f"=(1+{refs['ke_usd']})*(1+{p['inflacao_brl']})/(1+{p['inflacao_usd']})-1",
        PCT,
    )
    aba.pular()

    aba.secao("Custo da divida (Kd)")
    if "custo_divida_brl" in p:
        refs["kd_bruto"] = aba.ligacao(
            "Kd bruto em BRL (informado)", f"={p['custo_divida_brl']}", PCT
        )
    else:
        refs["kd_usd"] = aba.formula(
            "Kd em USD nominal (sintetico)",
            f"={p['rf_usd']}+{p['risco_pais']}+{p['spread_credito']}",
            PCT,
        )
        refs["kd_bruto"] = aba.formula(
            "Kd bruto em BRL nominal",
            f"=(1+{refs['kd_usd']})*(1+{p['inflacao_brl']})/(1+{p['inflacao_usd']})-1",
            PCT,
        )
    refs["kd_liquido"] = aba.formula(
        "Kd apos IR",
        f"={refs['kd_bruto']}*(1-{p['aliquota_ir']})",
        PCT,
    )
    aba.pular()

    aba.secao("WACC")
    refs["peso_divida"] = aba.formula(
        "Peso da divida", f"={p['divida_pl_alvo']}/(1+{p['divida_pl_alvo']})", PCT
    )
    refs["peso_equity"] = aba.formula("Peso do equity", f"=1-{refs['peso_divida']}", PCT)
    refs["wacc"] = aba.total(
        "WACC (BRL nominal)",
        f"={refs['peso_equity']}*{refs['ke_brl']}+{refs['peso_divida']}*{refs['kd_liquido']}",
        PCT,
    )
    aba.pular()
    aba.nota(
        f"Conferencia (calculado no Python na geracao): WACC = "
        f"{resultado.custo_capital.wacc_brl:.2%} | Ke = {resultado.custo_capital.ke_brl:.2%}"
    )
    if cc.lambda_pais != 1.0:
        aba.nota(f"Lambda de {cc.lambda_pais:.2f} aplicado ao premio de risco-pais.")

    aba.ajustar(largura_rotulo=42, n=4)
    return refs


def _aba_projecao(
    wb: Workbook, resultado: ResultadoValuation, p: dict
) -> dict[str, dict[str, str]]:
    proj = resultado.projecao
    n = proj.horizonte
    aba = _Aba(wb.create_sheet("Projecao"))

    aba.titulo(f"Projecao de fluxo de caixa - {resultado.empresa.nome}", largura=n + 3)
    aba.nota(f"Valores em {resultado.empresa.unidade}. Todas as linhas sao formulas.")
    aba.pular()
    aba.cabecalho_anos(proj.anos)

    def col(chave: str, i: int) -> str:
        return p[chave][str(i)]

    # A receita e uma linha encadeada: cada ano parte do anterior. Como a formula
    # precisa apontar para a propria linha, o endereco e calculado antes da escrita.
    linha_receita = aba.linha
    anterior_receita = [
        f"'{aba.ws.title}'!${get_column_letter(2 + i)}${linha_receita}" for i in range(n)
    ]
    receita = aba.linha_anual(
        "Receita liquida",
        [f"={anterior_receita[i]}*(1+{col('crescimento', i)})" for i in range(n)],
        MOEDA,
        base=f"={p['receita_base']}",
    )

    ebitda = aba.linha_anual(
        "EBITDA",
        [f"={receita[str(i)]}*{col('margem_ebitda', i)}" for i in range(n)],
        MOEDA,
    )
    depreciacao = aba.linha_anual(
        "(-) Depreciacao e amortizacao",
        [f"=-{receita[str(i)]}*{col('dep_pct', i)}" for i in range(n)],
        MOEDA,
    )
    ebit = aba.linha_anual(
        "EBIT",
        [f"={ebitda[str(i)]}+{depreciacao[str(i)]}" for i in range(n)],
        MOEDA,
        negrito=True,
    )
    impostos = aba.linha_anual(
        "(-) IR/CSLL sobre EBIT",
        [f"=-MAX({ebit[str(i)]},0)*{p['aliquota_ir']}" for i in range(n)],
        MOEDA,
    )
    nopat = aba.linha_anual(
        "NOPAT",
        [f"={ebit[str(i)]}+{impostos[str(i)]}" for i in range(n)],
        MOEDA,
        negrito=True,
    )
    dep_volta = aba.linha_anual(
        "(+) Depreciacao e amortizacao",
        [f"=-{depreciacao[str(i)]}" for i in range(n)],
        MOEDA,
    )
    capex = aba.linha_anual(
        "(-) Capex",
        [f"=-{receita[str(i)]}*{col('capex_pct', i)}" for i in range(n)],
        MOEDA,
    )
    aba.pular()
    giro = aba.linha_anual(
        "Capital de giro liquido (saldo)",
        [f"={receita[str(i)]}*{col('giro_pct', i)}" for i in range(n)],
        MOEDA,
        base=f"={p['receita_base']}*{col('giro_pct', 0)}",
    )
    var_giro = aba.linha_anual(
        "(-) Variacao do capital de giro",
        [
            f"=-({giro[str(i)]}-{giro['base'] if i == 0 else giro[str(i - 1)]})"
            for i in range(n)
        ],
        MOEDA,
    )
    aba.pular()
    fcff = aba.linha_anual(
        "FCFF (fluxo para a firma)",
        [
            f"={nopat[str(i)]}+{dep_volta[str(i)]}+{capex[str(i)]}+{var_giro[str(i)]}"
            for i in range(n)
        ],
        MOEDA,
        negrito=True,
    )

    aba.pular()
    aba.secao("Indicadores de consistencia", largura=n + 3)
    aba.linha_anual(
        "Margem EBITDA",
        [f"={ebitda[str(i)]}/{receita[str(i)]}" for i in range(n)],
        PCT1,
    )
    aba.linha_anual(
        "Margem EBIT", [f"={ebit[str(i)]}/{receita[str(i)]}" for i in range(n)], PCT1
    )
    aba.linha_anual(
        "Capex / Depreciacao",
        [f"=IF({dep_volta[str(i)]}=0,\"n/a\",-{capex[str(i)]}/{dep_volta[str(i)]})" for i in range(n)],
        MULTIPLO,
    )
    aba.nota(
        "Capex/Depreciacao muito abaixo de 1x no ultimo ano indica que a empresa "
        "entra na perpetuidade sem repor o ativo imobilizado."
    )

    aba.ajustar(n=n)
    return {"receita": receita, "ebitda": ebitda, "nopat": nopat, "fcff": fcff}


def _aba_dcf(
    wb: Workbook,
    resultado: ResultadoValuation,
    p: dict,
    cc: dict,
    proj: dict,
) -> None:
    dcf = resultado.dcf
    n = resultado.projecao.horizonte
    aba = _Aba(wb.create_sheet("DCF"))
    perp = resultado.empresa.perpetuidade

    aba.titulo(f"Fluxo de caixa descontado - {resultado.empresa.nome}", largura=n + 3)
    convencao = "meio de ano (t - 0,5)" if dcf.meio_de_ano else "fim de ano (t)"
    aba.nota(f"Convencao de desconto: {convencao}.")
    aba.pular()

    taxa = aba.ligacao("Taxa de desconto (WACC)", f"={cc['wacc']}", PCT)
    aba.pular()

    aba.cabecalho_anos(resultado.projecao.anos, coluna_base=False)
    periodo = aba.linha_anual(
        "Periodo de desconto",
        [float(i + 1) - (0.5 if dcf.meio_de_ano else 0.0) for i in range(n)],
        NUMERO,
    )
    fluxo = aba.linha_anual(
        "FCFF", [f"={proj['fcff'][str(i)]}" for i in range(n)], MOEDA
    )
    fator = aba.linha_anual(
        "Fator de desconto",
        [f"=1/(1+{taxa})^{periodo[str(i)]}" for i in range(n)],
        "0.0000",
    )
    descontado = aba.linha_anual(
        "Fluxo descontado",
        [f"={fluxo[str(i)]}*{fator[str(i)]}" for i in range(n)],
        MOEDA,
        negrito=True,
    )
    aba.pular()

    faixa = f"{_celula(descontado['0'])}:{_celula(descontado[str(n - 1)])}"
    vp_explicito = aba.total(
        "VP dos fluxos do periodo explicito", f"=SUM({faixa})", MOEDA
    )
    aba.pular()

    aba.secao("Valor terminal")
    if perp.metodo == "gordon":
        if perp.roic_perpetuidade is not None:
            expressao = (
                f"={proj['nopat'][str(n - 1)]}*(1+{p['g_perpetuo']})"
                f"*(1-{p['g_perpetuo']}/{p['roic_perp']})/({taxa}-{p['g_perpetuo']})"
            )
            aba.nota(
                "Gordon com reinvestimento normalizado: o fluxo perpetuo desconta "
                "a taxa de reinvestimento g/ROIC exigida para sustentar o crescimento."
            )
        else:
            expressao = (
                f"={proj['fcff'][str(n - 1)]}*(1+{p['g_perpetuo']})"
                f"/({taxa}-{p['g_perpetuo']})"
            )
            aba.nota("Gordon simples sobre o FCFF do ultimo ano projetado.")
        valor_terminal = aba.formula("Valor terminal (fim do ano n)", expressao, MOEDA)
    else:
        valor_terminal = aba.formula(
            "Valor terminal (fim do ano n)",
            f"={proj['ebitda'][str(n - 1)]}*{p['multiplo_saida']}",
            MOEDA,
        )
        aba.nota("Multiplo de saida aplicado sobre o EBITDA do ultimo ano projetado.")

    vp_terminal = aba.formula(
        "VP do valor terminal", f"={valor_terminal}/(1+{taxa})^{n}", MOEDA
    )
    aba.pular()

    ev = aba.total("Enterprise Value", f"={vp_explicito}+{vp_terminal}", MOEDA)
    peso = aba.formula("% do EV vindo da perpetuidade", f"={vp_terminal}/{ev}", PCT1)
    aba.nota(
        "Acima de ~75% costuma indicar projecao explicita curta demais: quase todo "
        "o valor passa a depender das premissas de perpetuidade."
    )
    aba.pular()

    aba.secao("Ponte para o Equity Value")
    itens = [
        ("(-) Divida bruta", f"=-{p['divida_bruta']}"),
        ("(+) Caixa e equivalentes", f"={p['caixa']}"),
        ("(+) Aplicacoes financeiras", f"={p['aplicacoes']}"),
        ("(-) Participacao de minoritarios", f"=-{p['minoritarios']}"),
        ("(-) Contingencias", f"=-{p['contingencias']}"),
        ("(-) Deficit atuarial", f"=-{p['deficit_atuarial']}"),
        ("(+) Ativos nao operacionais", f"={p['ativos_nao_op']}"),
    ]
    refs_ponte = [aba.formula(rotulo, expr, MOEDA) for rotulo, expr in itens]
    soma_ponte = "+".join(refs_ponte)
    equity = aba.total("Equity Value", f"={ev}+{soma_ponte}", MOEDA)
    aba.formula(
        "Valor por acao",
        f'=IF(N({p["acoes"]})=0,"informe as acoes em circulacao",{equity}/{p["acoes"]})',
        MOEDA,
    )

    aba.pular()
    aba.nota(
        f"Conferencia (Python): EV = {dcf.enterprise_value:,.1f} | "
        f"Equity = {dcf.equity_value:,.1f}"
        + (f" | por acao = {dcf.valor_por_acao:,.2f}" if dcf.valor_por_acao else "")
    )
    aba.ajustar(largura_rotulo=40, n=n)


def _aba_retorno(wb: Workbook, resultado, retorno, acionista) -> None:
    """Aba do TSR, com formulas vivas como as demais abas do modelo.

    A TIR sai da funcao IRR do proprio Excel, sobre a mesma linha de fluxos que
    o leitor ve; as contribuicoes sao formulas sobre as celulas de entrada. Quem
    receber o arquivo muda o preco de entrada ou o multiplo de saida e ve o
    retorno se refazer -- que e exatamente o uso desta analise.
    """
    n = retorno.anos
    aba = _Aba(wb.create_sheet("Retorno (TSR)"))

    aba.titulo(f"Retorno esperado do acionista - {resultado.empresa.nome}", largura=n + 3)
    aba.nota(
        "TSR = crescimento do lucro + dividendos + variacao de multiplo + termo "
        "cruzado. As parcelas somam o total exatamente."
    )
    aba.pular()

    aba.secao("Entrada e saida")
    preco = aba.entrada("Preco pago pelo equity", retorno.preco_entrada, MOEDA)
    lucro_entrada = aba.entrada("Lucro liquido de entrada", retorno.lucro_entrada, MOEDA)
    multiplo_entrada = aba.formula("P/L de entrada", f"={preco}/{lucro_entrada}", MULTIPLO)
    lucro_saida = aba.entrada("Lucro liquido de saida", retorno.lucro_saida, MOEDA)
    multiplo_saida = aba.entrada("P/L de saida", retorno.multiplo_saida, MULTIPLO)
    preco_saida = aba.formula("Preco de saida", f"={lucro_saida}*{multiplo_saida}", MOEDA)
    anos = aba.entrada("Anos ate a saida", float(n), "0")
    aba.pular()

    aba.secao("Fluxos do investidor", largura=n + 3)
    aba.cabecalho_anos(list(range(1, n + 1)))
    dividendos = aba.linha_anual(
        "Dividendos recebidos",
        [float(v) for v in retorno.dividendos[:n]],
        MOEDA,
        base=0.0,
        fonte=AZUL_ENTRADA,
    )
    venda = aba.linha_anual(
        "Venda da participacao",
        [0.0 if i < n - 1 else f"={preco_saida}" for i in range(n)],
        MOEDA,
        base=0.0,
    )
    # A coluna "Base" carrega o desembolso da compra, para que a faixa passada a
    # IRR comece no ano zero: sem o fluxo negativo dentro da propria faixa, a
    # funcao do Excel nao tem raiz para encontrar.
    fluxo = aba.linha_anual(
        "Fluxo do investidor",
        [f"={dividendos[str(i)]}+{venda[str(i)]}" for i in range(n)],
        MOEDA,
        base=f"=-{preco}",
        negrito=True,
    )
    aba.pular()

    faixa = f"{_celula(fluxo['base'])}:{_celula(fluxo[str(n - 1)])}"
    aba.secao("Decomposicao do retorno")
    tsr = aba.total("TSR esperado (TIR)", f"=IRR({faixa})", PCT)
    crescimento = aba.formula(
        "Crescimento do lucro", f"=({lucro_saida}/{lucro_entrada})^(1/{anos})-1", PCT
    )
    multiplo = aba.formula(
        "Variacao de multiplo", f"=({multiplo_saida}/{multiplo_entrada})^(1/{anos})-1", PCT
    )
    cruzado = aba.formula("Termo cruzado", f"={crescimento}*{multiplo}", PCT)
    dividendo_residual = aba.formula(
        "Dividendos", f"={tsr}-(1+{crescimento})*(1+{multiplo})+1", PCT
    )
    aba.formula(
        "Soma das parcelas (confere com o TSR)",
        f"={crescimento}+{multiplo}+{cruzado}+{dividendo_residual}",
        PCT,
    )
    aba.pular()

    aba.nota(
        f"Conferencia (Python): TSR = {retorno.tsr:.2%} | crescimento = "
        f"{retorno.contribuicao_crescimento:.2%} | dividendos = "
        f"{retorno.contribuicao_dividendos:.2%} | multiplo = "
        f"{retorno.contribuicao_multiplo:.2%}"
    )
    if resultado.dcf.meio_de_ano:
        aba.nota(
            "Atencao: o modelo usa convencao de meio de ano, e a funcao IRR do Excel "
            "posiciona os fluxos no fim de cada ano. Por isso a celula acima fica "
            "levemente abaixo da conferencia em Python, que honra a convencao."
        )

    if acionista is not None:
        aba.pular()
        aba.secao("Da operacao ao bolso do acionista", largura=n + 3)
        aba.cabecalho_anos(acionista.anos, coluna_base=False)
        for rotulo, linha in acionista.tabela().iterrows():
            aba.linha_anual(str(rotulo), [float(v) for v in linha], MOEDA)

    aba.ajustar(largura_rotulo=36, n=n)


def _escrever_dataframe(
    aba: _Aba,
    df: pd.DataFrame,
    formato: str = MOEDA,
    formato_indice: str = "General",
    largura: int = 10,
) -> tuple[int, int]:
    """Escreve um DataFrame e devolve (linha do cabecalho, primeira linha de dados)."""
    linha_cabecalho = aba.linha
    aba.ws.cell(row=linha_cabecalho, column=1, value=str(df.index.name or ""))
    for j, coluna in enumerate(df.columns):
        celula = aba.ws.cell(row=linha_cabecalho, column=2 + j, value=str(coluna))
        celula.font = Font(bold=True)
        celula.fill = FUNDO_SECAO
        celula.alignment = Alignment(horizontal="center")
    aba.linha += 1
    primeira = aba.linha

    for indice, linha_dados in df.iterrows():
        celula = aba.ws.cell(row=aba.linha, column=1, value=str(indice))
        celula.font = Font(bold=True)
        celula.number_format = formato_indice
        for j, valor in enumerate(linha_dados):
            destino = aba.ws.cell(row=aba.linha, column=2 + j)
            if isinstance(valor, (int, float, np.floating, np.integer)):
                valor = float(valor)
                destino.value = None if not np.isfinite(valor) else valor
                destino.number_format = formato
                if destino.value is None:
                    destino.value = "n/a"
            else:
                destino.value = str(valor)
        aba.linha += 1

    aba.ajustar(largura_rotulo=28, largura_dados=largura, n=len(df.columns) + 1)
    return linha_cabecalho, primeira


def _aba_multiplos(wb: Workbook, comparaveis: list[Comparavel], alvo: Alvo | None) -> None:
    aba = _Aba(wb.create_sheet("Multiplos"))
    aba.titulo("Avaliacao relativa por comparaveis")
    aba.nota(
        "Multiplos calculados no Python na geracao do arquivo. "
        "Denominador nao positivo (EBITDA ou lucro negativo) aparece como n/a."
    )
    aba.pular()

    aba.secao("Multiplos dos comparaveis")
    _escrever_dataframe(aba, tabela_comparaveis(comparaveis), MULTIPLO)
    aba.pular()

    aba.secao("Estatisticas do peer group")
    _escrever_dataframe(aba, estatisticas(comparaveis), MULTIPLO)
    aba.pular()

    if alvo is not None:
        aba.secao(f"Valor implicito de {alvo.nome} (mediana do setor)")
        tabela = avaliar_por_multiplos(alvo, comparaveis)
        aba.ws.cell(row=aba.linha, column=1, value="Multiplo").font = Font(bold=True)
        for j, coluna in enumerate(tabela.columns):
            celula = aba.ws.cell(row=aba.linha, column=2 + j, value=str(coluna))
            celula.font = Font(bold=True)
            celula.fill = FUNDO_SECAO
        aba.linha += 1
        for indice, linha_dados in tabela.iterrows():
            aba.ws.cell(row=aba.linha, column=1, value=str(indice)).font = Font(bold=True)
            for j, (nome_coluna, valor) in enumerate(linha_dados.items()):
                celula = aba.ws.cell(row=aba.linha, column=2 + j)
                valor = float(valor)
                if not np.isfinite(valor):
                    celula.value = "n/a"
                else:
                    celula.value = valor
                    celula.number_format = (
                        MULTIPLO if "setor" in str(nome_coluna) else MOEDA
                    )
            aba.linha += 1
        aba.ajustar(largura_rotulo=18, largura_dados=18, n=6)


def _aba_sensibilidade(wb: Workbook, tabela: pd.DataFrame) -> None:
    aba = _Aba(wb.create_sheet("Sensibilidade"))
    aba.titulo("Sensibilidade bidimensional")
    aba.nota(
        f"Linhas: {tabela.index.name} | Colunas: {tabela.columns.name}. "
        "Valores calculados no Python; nao recalculam ao editar as premissas."
    )
    aba.pular()
    cabecalho, primeira = _escrever_dataframe(aba, tabela, MOEDA, largura=14)

    ultima = primeira + len(tabela) - 1
    ultima_coluna = get_column_letter(1 + len(tabela.columns))
    faixa = f"B{primeira}:{ultima_coluna}{ultima}"
    aba.ws.conditional_formatting.add(
        faixa,
        ColorScaleRule(
            start_type="min", start_color="F8696B",
            mid_type="percentile", mid_value=50, mid_color="FFEB84",
            end_type="max", end_color="63BE7B",
        ),
    )


def _aba_cenarios(wb: Workbook, tabela: pd.DataFrame) -> None:
    aba = _Aba(wb.create_sheet("Cenarios"))
    aba.titulo("Cenarios")
    aba.nota("Conjuntos coerentes de premissas, avaliados no Python.")
    aba.pular()
    _escrever_dataframe(aba, tabela, MOEDA, largura=16)


def _aba_monte_carlo(wb: Workbook, simulacao: ResultadoSimulacao) -> None:
    aba = _Aba(wb.create_sheet("Monte Carlo"))
    aba.titulo("Simulacao de Monte Carlo")
    aba.nota(
        f"Metrica simulada: {simulacao.metrica} | {simulacao.simulacoes:,} rodadas | "
        f"{simulacao.descartadas:,} descartadas por inviabilidade economica."
    )
    aba.pular()

    aba.secao("Estatisticas")
    resumo = simulacao.resumo()
    for rotulo, valor in resumo.items():
        aba.ws.cell(row=aba.linha, column=1, value=rotulo).font = Font(bold=True)
        celula = aba.ws.cell(row=aba.linha, column=2, value=float(valor))
        celula.number_format = MOEDA
        aba.linha += 1
    aba.pular()

    aba.secao("Distribuicao")
    contagens, bordas = np.histogram(simulacao.valores, bins=30)
    linha_inicial = aba.linha
    aba.ws.cell(row=aba.linha, column=1, value="Faixa").font = Font(bold=True)
    aba.ws.cell(row=aba.linha, column=2, value="Frequencia").font = Font(bold=True)
    aba.linha += 1
    primeira_dados = aba.linha
    for i, contagem in enumerate(contagens):
        centro = (bordas[i] + bordas[i + 1]) / 2
        aba.ws.cell(row=aba.linha, column=1, value=float(centro)).number_format = MOEDA
        aba.ws.cell(row=aba.linha, column=2, value=int(contagem))
        aba.linha += 1
    ultima_dados = aba.linha - 1

    grafico = BarChart()
    grafico.title = f"Distribuicao simulada - {simulacao.metrica}"
    grafico.y_axis.title = "Frequencia"
    grafico.x_axis.title = simulacao.metrica
    grafico.gapWidth = 8
    dados = Reference(aba.ws, min_col=2, min_row=linha_inicial, max_row=ultima_dados)
    categorias = Reference(aba.ws, min_col=1, min_row=primeira_dados, max_row=ultima_dados)
    grafico.add_data(dados, titles_from_data=True)
    grafico.set_categories(categorias)
    grafico.height, grafico.width = 9, 20
    aba.ws.add_chart(grafico, f"D{linha_inicial}")

    aba.ajustar(largura_rotulo=26, largura_dados=16, n=3)


def _aba_leia_me(wb: Workbook, resultado: ResultadoValuation) -> None:
    aba = _Aba(wb.create_sheet("Leia-me", 0))
    aba.titulo("Como usar esta planilha")
    aba.pular()

    linhas = [
        ("Convencao de cores", ""),
        ("  Azul", "premissa editavel - e aqui que voce mexe"),
        ("  Preto", "formula calculada dentro da propria aba"),
        ("  Verde", "referencia trazida de outra aba"),
        ("", ""),
        ("Abas com formulas vivas", "Premissas, Custo de Capital, Projecao, DCF"),
        (
            "Abas com valores fixos",
            "Multiplos, Sensibilidade, Cenarios, Monte Carlo - recalculadas so no Python",
        ),
        ("", ""),
        ("Empresa", resultado.empresa.nome),
        ("Data-base", resultado.empresa.data_base or "nao informada"),
        ("Moeda / unidade", f"{resultado.empresa.moeda} / {resultado.empresa.unidade}"),
        ("Horizonte explicito", f"{resultado.projecao.horizonte} anos"),
        ("", ""),
        ("WACC", f"{resultado.custo_capital.wacc_brl:.2%}"),
        ("Enterprise Value", f"{resultado.enterprise_value:,.1f}"),
        ("Equity Value", f"{resultado.equity_value:,.1f}"),
        ("% do EV na perpetuidade", f"{resultado.dcf.peso_perpetuidade:.1%}"),
    ]
    for rotulo, valor in linhas:
        aba.ws.cell(row=aba.linha, column=1, value=rotulo).font = Font(
            bold=not rotulo.startswith("  ")
        )
        aba.ws.cell(row=aba.linha, column=2, value=valor)
        aba.linha += 1

    aba.pular()
    aba.nota(
        "Este modelo e uma ferramenta de apoio. Os numeros valem o que valem as "
        "premissas: revise-as antes de usar o resultado em qualquer decisao."
    )
    aba.ajustar(largura_rotulo=26, largura_dados=70, n=2)

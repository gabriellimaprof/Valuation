"""Geracao do template de demonstracoes financeiras para preenchimento manual.

Quando o analista nao tem um export pronto, este e o caminho mais rapido: uma
planilha com as contas ja nomeadas, os anos nas colunas e uma aba de instrucoes.
Preenchida, ela e reimportada pelo mesmo ``importar`` das demais origens.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .esquema import CONTAS_BP, CONTAS_DFC, CONTAS_DRE

FUNDO_TITULO = PatternFill("solid", fgColor="1F3864")
FUNDO_SECAO = PatternFill("solid", fgColor="D9E1F2")
FUNDO_ENTRADA = PatternFill("solid", fgColor="FFFDE7")


def gerar_template(
    caminho: str | Path,
    anos: list[int] | None = None,
    empresa: str = "Nome da empresa",
) -> Path:
    """Grava o template vazio de demonstracoes financeiras."""
    caminho = Path(caminho)
    anos = anos or list(range(2020, 2026))

    wb = Workbook()
    _aba_instrucoes(wb.active)
    _aba_demonstracoes(wb.create_sheet("Demonstracoes"), anos, empresa)

    caminho.parent.mkdir(parents=True, exist_ok=True)
    wb.save(caminho)
    return caminho


def _aba_instrucoes(ws) -> None:
    ws.title = "Instrucoes"
    linhas = [
        ("Template de demonstracoes financeiras", True),
        ("", False),
        ("Preencha a aba 'Demonstracoes' e importe o arquivo de volta no app.", False),
        ("", False),
        ("Regras que evitam erro de importacao:", True),
        ("1. Nao renomeie as contas da coluna A. E por elas que o app reconhece as linhas.", False),
        ("2. Nao mude a linha dos anos. Voce pode acrescentar ou remover colunas de ano.", False),
        ("3. Use a mesma unidade em toda a planilha (por exemplo, R$ mil ou R$ milhoes).", False),
        ("4. Custos, despesas, impostos e capex podem ser positivos ou negativos:", False),
        ("   o app usa a magnitude e padroniza o sinal sozinho.", False),
        ("5. Deixe em branco o que nao tiver. O app deriva o que der para derivar", False),
        ("   (por exemplo, EBIT a partir de lucro bruto menos despesas operacionais).", False),
        ("", False),
        ("Contas minimas para o modelo rodar:", True),
        ("Receita liquida, EBIT, Lucro liquido, Ativo total, Patrimonio liquido,", False),
        ("Caixa e equivalentes.", False),
        ("", False),
        ("Quanto mais anos, melhor: a analise historica precisa de pelo menos 3 anos", False),
        ("para calcular crescimento, reinvestimento e ROIC com algum significado.", False),
    ]
    for indice, (texto, negrito) in enumerate(linhas, start=1):
        celula = ws.cell(row=indice, column=1, value=texto)
        celula.font = Font(bold=negrito, size=13 if indice == 1 else 11)
    ws.column_dimensions["A"].width = 95


def _aba_demonstracoes(ws, anos: list[int], empresa: str) -> None:
    ws.cell(row=1, column=1, value="Empresa").font = Font(bold=True)
    ws.cell(row=1, column=2, value=empresa).font = Font(color="0000CC", bold=True)
    ws.cell(row=2, column=1, value="Unidade").font = Font(bold=True)
    ws.cell(row=2, column=2, value="R$ mil").font = Font(color="0000CC", bold=True)

    linha = 4
    ws.cell(row=linha, column=1, value="Conta").font = Font(bold=True)
    for i, ano in enumerate(anos):
        celula = ws.cell(row=linha, column=2 + i, value=ano)
        celula.font = Font(bold=True)
        celula.fill = FUNDO_SECAO
        celula.alignment = Alignment(horizontal="center")
    linha += 1

    for titulo, contas in (
        ("DEMONSTRACAO DO RESULTADO", CONTAS_DRE),
        ("BALANCO PATRIMONIAL", CONTAS_BP),
        ("FLUXO DE CAIXA", CONTAS_DFC),
    ):
        linha += 1
        celula = ws.cell(row=linha, column=1, value=titulo)
        celula.font = Font(bold=True, color="FFFFFF")
        celula.fill = FUNDO_TITULO
        for coluna in range(2, len(anos) + 2):
            ws.cell(row=linha, column=coluna).fill = FUNDO_TITULO
        linha += 1

        for conta in contas:
            rotulo = ws.cell(row=linha, column=1, value=conta.rotulo)
            if conta.obrigatoria:
                rotulo.font = Font(bold=True)
            if conta.ajuda:
                rotulo.comment = None  # mantido simples: a ajuda vai na coluna final
            for coluna in range(2, len(anos) + 2):
                celula = ws.cell(row=linha, column=coluna)
                celula.fill = FUNDO_ENTRADA
                celula.number_format = "#,##0.0"
            ws.cell(
                row=linha,
                column=len(anos) + 3,
                value=("[obrigatoria] " if conta.obrigatoria else "") + conta.ajuda,
            ).font = Font(italic=True, size=9, color="777777")
            linha += 1

    ws.column_dimensions["A"].width = 44
    for coluna in range(2, len(anos) + 2):
        ws.column_dimensions[get_column_letter(coluna)].width = 14
    ws.column_dimensions[get_column_letter(len(anos) + 3)].width = 80
    ws.freeze_panes = ws.cell(row=5, column=2)
